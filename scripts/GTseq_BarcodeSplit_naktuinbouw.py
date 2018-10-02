#!/usr/bin/env python3
"""
GTseq_BarcodeSplit_MP.py
by Nate Campbell
edit: Arlo Hoogeveen 2018, added options and is now compatible with excel template.
Use this script to split out individual files by barcode combination rather than using "grep method".
This version will process up to 10,500 individual samples and simultaneously run up to 21 processing cores for faster parsing of individual sequences.
Input file 1 is a .csv file containing sample and barcode information [individual sample names, PlateID,i7_name,i7_sequence,i5_name,i5_sequence] for each sample.
example: Sample,PlateID,i7_name,i7_sequence,i5_name,i5_sequence
         Sample123,P1234,i001,ACCGTA,25,CCCTAA
         Sample321,P1234,i001,ACCGTA,26,GGCACA
         ....
Note: The header line is ignored while executing the script so always include it to avoid missing data.  Also, note that output files are appended
and therefore not overwritten if the script is called multiple times.  If for some reason the script needs to be run more than once for the same set of samples,
the original output files will need to be deleted to avoid having files with duplicated sequences.
Input file 2 is the .fastq file containing the sequences.
Output is a set of fastq files for all individuals included in the library named in the format [i7name]_[i5name]_[PlateID]_[SampleID].fastq
This script uses multiple processors (1 for every 500 samples) and was clocked at about 20 minutes to complete barcode splitting for one lane of data.
Speed is faster than grep method while using fewer compute resources.
Using this script for barcode splitting and the GTseq_Genotyper_v2.pl script for genotyping allows the GTseq pipeline to be run on a linux
desktop computer.  (Raw data to genotypes in less than 1 hour).
"""

from openpyxl import load_workbook
import os
import argparse


def argument_parser():
     parser = argparse.ArgumentParser(description='Options for BarcodeSplit script')
     parser.add_argument(
        "-i","--input",
        dest="barcode_file",
        help="path to the .csv file containing sample and barcode information"
     )
     parser.add_argument(
        "-f","--fastq",
        dest="sequence_file",
        help="path to the .fastq file containing the sequences"
     )
     parser.add_argument(
        "-S","--sequencer",
        dest="sequencer_id",
        help="the sequencer id that the info line starts with."
     )
     parser.add_argument(
         "-d","--dir",
         default=os.getcwd(),
         help="the directory to put the files into."
     )
     parser.add_argument(
         "-s","--sep",
         default = "+",
         help="the charector seperating the i7 and i5 barcode sequence in the fastq."
     )
     return parser.parse_args()


def split_file(individual_list, info_line_start, path2, separator='+'):

     fq = open(path2, 'r')
     handle_dict = {}

     for bar,name in individual_list.items():
          handle_dict[bar] = open(name, 'a+')

     lineNo2 = 0
     writelines = 0

     for line in fq:
          lineNo2 = lineNo2 + 1
          if writelines < 5 and writelines > 0:
               f_out.write(line)
               writelines = writelines + 1
               if writelines == 4:
                    writelines = 0

          if line.startswith(info_line_start):
               info = line.split(':')
               BC = '+'.join([bc[:6] for bc in info[9].split(separator)])
               if BC in handle_dict:
                    f_out = handle_dict[BC]
                    f_out.write(line)
                    writelines = 1
     fq.close()
     return

def fetch_barcode_info(infile,folder):
    """
    Make sample list with the excel template.
    """
    #names as created in GTseq_BarcodeSplit
    filename_parts = ['i7_name', 'i5_name', 'Plate_name']
    barcode_parts = ['i7_barcode', 'i5_barcode']
    filenames = {}
    wb = load_workbook(infile, data_only=True)
    info_sheet = wb['Library']
    info_headers = []
    for column in info_sheet.iter_cols(max_row=1):
        info_headers.append(column[0].value)
    for row in info_sheet.iter_rows(min_row=2):
        info_dict = {info_headers[i]:item.value for i,item in enumerate(row)}
        # makeing sure the folders exist
        if not os.path.exists(folder):
            os.makedirs(folder)
        # rows with empty cells are not allowed
        if all(x for x in info_dict.values()):
            species_folder = os.path.join(folder, info_dict['Species'].lower())
            if not os.path.exists(species_folder):
                os.makedirs(species_folder)
            filename = '_'.join(
                ['_'.join(info_dict[x].split()) for x in filename_parts]
            )+'.fastq'
            barcode = '+'.join([info_dict[x] for x in barcode_parts])
            filenames[barcode] = os.path.join(species_folder, filename)
    return [filenames]


def get_barcode_and_filename(line,folder):
    stuff = line.strip().split(',')
    if not os.path.exists(folder):
        os.makedirs(folder)
    species_folder = os.path.join(folder, stuff[0])
    if not os.path.exists(species_folder):
        os.makedirs(species_folder)
    filename = os.path.join(species_folder, (stuff[2] + '_' + stuff[4] + '_' + stuff[1] + '.fastq'))
    barcode = stuff[3] + '+' + stuff[5]
    return barcode,filename


def make_sample_list(path1,folder):
    lists = []

    max_pool_size = 500
    iterator = 0

    with open(path1,'r') as infile:
        headers = infile.readline()
        list = {}
        for line in infile:
            iterator += 1
            if iterator%max_pool_size==0:
                lists.append(list)
                list = {}
            #get the filename and barcode,
            barcode, filename = get_barcode_and_filename(line,folder)
            list[barcode] = filename
        if list:
            lists.append(list)
    return lists


def main(args):
    # ask for barcode file if not given.
    if args.barcode_file:
        path1 = args.barcode_file
    else:
        print('type the path to input file\nFormat= /home/user/...')
        path1 = input()

    # ask for sequence file if not given.
    if args.sequence_file:
        path2 = args.sequence_file
    else:
        print('type the path to the fastq file to split\nFormat= /home/user/...')
        path2 = input()

    # getting the string the headers start with
    if args.sequencer_id:
        info_line_start = args.sequencer_id
    else:
        # using the first line to grab the info line indication
        with open(path2, 'r') as seq_file:
            info_line_start = seq_file.readline().split(':')[0]

    #limiting the amount of file connections to 500.
    if path1.split('.')[-1] == 'csv':
        lists = make_sample_list(path1,args.dir)
    elif path1.split('.')[-1] == 'xlsm':
        lists = fetch_barcode_info(path1,args.dir)
    else:
        lists=[]
        print('infofile with not supported extension')
    for list in lists:
        split_file(list,info_line_start,path2,args.sep)

if __name__ == '__main__':
    args = argument_parser()
    main(args)