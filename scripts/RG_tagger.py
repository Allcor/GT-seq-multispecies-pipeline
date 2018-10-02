#!/usr/bin/python3
"""
http://gitlab.naktuinbouw.net/bioinformatics/gt-seq/issues/18

adding a Sam RG tag to each read containing the flowcell ID, flowcell lane and Sample ID
 `RG:Z:C5M8DANXX_3_Mini_Stars_2`

Input is
 1: fastq file coming from a GT_seq run
 2: xlsx file with barcodes and corresponding sample ID's

 Output is
 1: fastq file with the tags now attached.
"""

from argparse import ArgumentParser
from openpyxl import load_workbook
import re
import unicodedata
import sys


def argument_parser():
    parser = ArgumentParser(description='Options for fastq RG tagger')
    parser.add_argument(
        '-f', '--fastq',
        required=True,
        help="the fastq file that will be receiving the tags"
    )
    parser.add_argument(
        '-i', '--info',
        required=True,
        help="the .csv or .xlsx file with information on the read groups"
    )
    parser.add_argument(
        '-o', '--fastq_out',
        help="the new fastq file with tags"
    )
    parser.add_argument(
        '-s', '--split',
        help="add this option if you want a fastq file for each read group",
        action='store_true'
    )
    parser.add_argument(
        '-u', '--fuzz',
        help="fuzzy matching of barcodes",
        action='store_true'
    )
    return parser.parse_args()


def add_header_tag(header_line,readgroups,return_barcode=False):
    bc_len = 6
    bc_split = '+'
    sequencer_dict, read_dict = obtain_read_info(header_line)
    barcode_ = [bc[:bc_len] for bc in read_dict['barcode'].strip().split(bc_split)]
    barcode = bc_split.join(barcode_)
    # check if the barcode is used and with how many mismatches
    mismatch = None
    RG_lib = None
    for i, readgroup_dict in enumerate(readgroups):
        if barcode in readgroup_dict:
            RG_lib = readgroup_dict[barcode]
            mismatch = i
    if RG_lib:
        if not RG_lib['Flowcell'] == sequencer_dict['flowcell_id']:
            print(header_line+" has different flowcell then expected form barcode file")
        if not str(RG_lib['Lane']) == sequencer_dict['flowcell_lane']:
            print(header_line+" has different lane then expected form barcode file")
        RG_tag = RG_lib['ID']
        header_line = header_line.split()[0]+' bl:Z:{}\tbr:Z:{}\tbm:Z:{}\tRG:Z:{}\n'.format(
            barcode_[0],
            barcode_[1],
            mismatch,
            RG_tag
        )
        barcode = RG_lib['barcode']
    else:
        header_line = header_line.split()[0]+' bl:Z:{}\tbr:Z:{}\n'.format(
            barcode_[0],
            barcode_[1],
        )
        barcode = None
    if return_barcode:
        return header_line, barcode
    else:
        return header_line


def remove_accents(input_str):
    """
    taken from https://stackoverflow.com/a/517974/3375944
    :param input_str: string with accented letters
    :return: string without accented letters
    """
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return u"".join([c for c in nfkd_form if not unicodedata.combining(c)])


def obtain_read_info(header_line):
    sequencer_stuff = header_line.split()[0]
    read_stuff = header_line.split()[1]
    # get_sequencer info from the read header
    sequencer_keys = [
        'instrument_name',
        'run_id',
        'flowcell_id',
        'flowcell_lane',
        'tile_nr',
        'cluster_x',
        'cluster_y'
    ]
    sequencer_dict = {sequencer_keys[i]:item for i,item in enumerate(sequencer_stuff.split(':'))}
    # get read info from the read header
    read_keys = [
        'pair',
        'filtered',
        'bit_flag',
        'barcode']
    read_dict = {read_keys[i]:item for i,item in enumerate(read_stuff.split(':'))}
    return sequencer_dict,read_dict


def obtain_barcode_info_xlsx(info_file):
    info_lines = []
    wb = load_workbook(info_file, data_only=True)
    if 'Library' in wb.sheetnames:
        info_sheet = wb['Library']
    else:
        info_sheet = wb.active
    info_headers = []
    for column in info_sheet.iter_cols(max_row=1):
        info_headers.append(column[0].value)
    for row in info_sheet.iter_rows(min_row=2):
        info_dict = {info_headers[i]:item.value for i,item in enumerate(row)}
        # rows with empty cells are not allowed
        if all(x for x in info_dict.values()):
            info_lines.append(info_dict)
        else:
            #check if library_info is filled in correctly.
            if info_dict['Sample_name']:
                if not info_dict['Flowcell'] or not info_dict['Lane']:
                    raise ValueError ("please fill in Flowcell and Lane from sequence run.")
    return info_lines


def obtain_barcode_info_csv(info_file):
    '''
    takes a csv file containing information of the sequencing library
    Puts each line in a dictionary with the headers as value
    each line is added to a list and this list is returned
    :param info_file:
    :return: list of dictionaries of samples
    '''
    # expected headers:
    # Flowcell, Lane, i7_barcode, i5_barcode, Species, Sample_name
    info_lines = []
    with open(info_file, 'r') as info:
        info_headers = re.split('[; ]+|[, ]+',info.readline().strip())
        for line in info:
            split_line = re.split('[; ]+|[, ]+',line.strip())
            info_dict = {info_headers[i]:item for i,item in enumerate(split_line)}
            info_lines.append(info_dict)
    return info_lines


def create_all_barcode_varients(barcode):
    varients = ['A','C','T','G']
    barcodes = ['N'+barcode[1:]]
    for i,base in enumerate(barcode):
        barcodes += [barcode[:i]+x+barcode[i+1:] for x in varients if x != base]
    return barcodes


def make_readgroups_dict(library_info_lines,make_var=False):
    '''
    reads the lines of the info file and returns a python dictionary
    :param library_info_lines: lines read from a data sheet
    :return: python dictionary with the lines
    '''
    readgroups_dict = {}
    readgroups_dict_1 = {}
    readgroups_dict_2 = {}
    for library_info in library_info_lines:
        readgroup = make_readgroup(library_info, clean_dict=False)
        #check if Sample_name used for ReadGroup_id is duplicated
        if readgroup['SM'] in [x['SM'] for x in readgroups_dict.values()]:
            raise ValueError ("Two samples have the same Sample_name, {}.".format(readgroup['SM']))
        readgroup['barcode'] = library_info['i7_barcode']+ '+' + library_info['i5_barcode']
        #creating all 1 base permutations so the readgroup can still be found.
        if make_var:
            i5_overlap_list = ['ACTCGT','AAACGT','AATCTT','AATCGG']
            i5_barcodes = create_all_barcode_varients(library_info['i5_barcode'])
            i5_barcodes = [bc for bc in i5_barcodes if (bc not in i5_overlap_list)]
            for i5_barcode in i5_barcodes:
                # one mismatch
                barcode = library_info['i7_barcode'] + '+' + i5_barcode
                #check if barcodes (and 1 base permutations) are unique
                if barcode in readgroups_dict or barcode in readgroups_dict_1 or barcode in readgroups_dict_2:
                    print(library_info)
                    raise ValueError("Two samples with the same barcode combination, {}.".format(barcode))
                readgroups_dict_1[barcode] = readgroup
            for i7_barcode in create_all_barcode_varients(library_info['i7_barcode']):
                #one mismatch
                barcode = i7_barcode + '+' + library_info['i5_barcode']
                #check if barcodes (and 1 base permutations) are unique
                if barcode in readgroups_dict or barcode in readgroups_dict_1 or barcode in readgroups_dict_2:
                    print(library_info)
                    raise ValueError("Two samples with the same barcode combination, {}.".format(barcode))
                readgroups_dict_1[barcode] = readgroup
                #both ends have 1 mismatch
                for i5_barcode in i5_barcodes:
                    barcode = i7_barcode + '+' + i5_barcode
                    #check if barcodes (and 1 base permutations) are unique
                    if barcode in readgroups_dict or barcode in readgroups_dict_1 or barcode in readgroups_dict_2:
                        print(library_info)
                        raise ValueError ("Two samples with the same barcode combination, {}.".format(barcode))
                    readgroups_dict_2[barcode] = readgroup
        barcode = library_info['i7_barcode'] + '+' + library_info['i5_barcode']
        #check if barcodes are unique
        if barcode in readgroups_dict:
            raise ValueError ("Two samples with the same barcode combination, {}.".format(barcode))
        readgroups_dict[barcode] = readgroup
    return readgroups_dict,readgroups_dict_1,readgroups_dict_2


def make_readgroup(library_info, clean_dict=True):
    '''
    takes a line, and substitutes the information needed for a readgroup.
    :param library_info: line containing relevant information on a sample
    :param clean_dict: will add readgroup to a clean dictionary instead of using the library_info input.
    :return: dictionary with info of readgroup
    '''
    #initiate returned library
    if clean_dict:
        readgroup_dict = {}
    else:
        readgroup_dict = library_info
    # remove problametic charecters.
    sample_name = remove_accents(str(library_info['Sample_name']))
    # ID
    # Read group identifier. Each @RG line must have a unique ID. The value of ID is used in the RG
    # tags of alignment records. Must be unique among all read groups in header section. Read group
    # IDs may be modified when merging SAM files in order to handle collisions.
    readgroup_dict['ID'] = '_'.join([
        library_info['Flowcell'],
        str(library_info['Lane']),
        sample_name
    ])
    # CN
    # Name of sequencing center producing the read.
    # DS
    # Description.
    # DT
    # Date the run was produced (ISO8601 date or date/time).
    # FO
    # Flow order. The array of nucleotide bases that correspond to the nucleotides used for each
    # flow of each read. Multi-base flows are encoded in IUPAC format, and non-nucleotide flows by
    # various other characters. Format: /\*|[ACMGRSVTWYHKDBN]+/
    # KS
    # The array of nucleotide bases that correspond to the key sequence of each read.
    readgroup_dict['KS'] = '+'.join([
        library_info['i7_barcode'],
        library_info['i5_barcode']
    ])
    # LB
    # Library.
    readgroup_dict['LB'] = '_'.join([
        library_info['Flowcell'],
        str(library_info['Lane'])
    ])
    # PG
    # Programs used for processing the read group.
    # PI
    # Predicted median insert size.
    # PL
    # Platform/technology used to produce the reads. Valid values: CAPILLARY, LS454, ILLUMINA,
    # SOLID, HELICOS, IONTORRENT, ONT, and PACBIO
    readgroup_dict['PL'] = 'ILLUMINA'
    # PM
    # Platform model. Free-form text providing further details of the platform/technology used.
    # PU
    # Platform unit (e.g. flowcell-barcode.lane for Illumina or slide for SOLiD). Unique identifier
    # SM
    # Sample. Use pool name where a pool is being sequenced.
    readgroup_dict['SM'] = sample_name
    return readgroup_dict


def main(args):
    # establish the readgroups
    info_filetype = args.info.split('.')[-1]
    if info_filetype == "csv":
        library_info_list = obtain_barcode_info_csv(args.info)
    elif info_filetype in ["xlsx", "xlsm"]:
        library_info_list = obtain_barcode_info_xlsx(args.info)
    else:
        raise("barcode file extention ."+info_filetype+" is not supported, needs to be .csv or .xlsx")
    readgroups,readgroups_1,readgroups_2 = make_readgroups_dict(library_info_list,args.fuzz)

    # groups used for split
    outfile_groups = {}
    for library_info in library_info_list:
        barcode = library_info['i7_barcode'] + '+' + library_info['i5_barcode']
        if 'Species' in library_info_list[0]:
            outfile_groups[barcode] = library_info['Species']
        else:
            outfile_groups[barcode] = 'main'

    # prepare the outfile
    outfiles = {}
    if args.split:
        if args.fastq_out:
            out_split = args.fastq_out.split('.')
            for group in set(outfile_groups.values()):
                group_outname = '.'.join(out_split[:-1])+"."+group.lower()+"."+out_split[-1]
                outfiles[group] = open(group_outname, 'w')
            other_outname = '.'.join(out_split[:-1])+".nogroup."+out_split[-1]
            outfiles['nogroup'] = open(other_outname, 'w')
        else:
            raise("can't have split withought specyfying the outfile location")
    else:
        if args.fastq_out:
            outfiles['main'] = open(args.fastq_out, 'w')
        else:
            outfiles['main'] = sys.stdout

    # read the entire fastq and fix the lines and write it in the new one
    with open(args.fastq, 'r') as fastqfile:
        i = 0
        read = {}
        outfile = sys.stdout
        for line in fastqfile:
            i += 1
            if i == 1:
                read['header_line'] = line
                # add the RG_tag
                if args.split:
                    # change the outfile if multiple files are requested
                    read['header_line_fix'], barcode = add_header_tag(
                        read['header_line'],
                        [readgroups,readgroups_1,readgroups_2],
                        args.split
                    )
                    if barcode:
                        outfile = outfiles[outfile_groups[barcode]]
                    else:
                        outfile = outfiles['nogroup']
                else:
                    read['header_line_fix'] = add_header_tag(
                        read['header_line'],
                        [readgroups,readgroups_1,readgroups_2]
                    )
                    outfile = outfiles['main']
            elif i == 2:
                read['sequence_line'] = line
            elif i == 3:
                read['spacer_line'] = line
            elif i == 4:
                read['quality_line'] = line
                # also write the line
                parts = ['header_line_fix', 'sequence_line', 'spacer_line', 'quality_line']
                outfile.writelines([read[part] for part in parts])
                i=0

    if args.fastq_out:
        # if the outfile was specified it needs to be closed
        if args.split:
            for group in set(outfile_groups.values()):
                outfiles[group].close()
        else:
            outfile.close()


if __name__ == '__main__':
    args = argument_parser()
    main(args)