#!/usr/local/bin/python3.6
"""
Script to add a filter flag to soft-clipping reads. This to prevent them from being used in mpileup.

Also gives a file with mapping stats for each read group.

Input
    BAM file,

output
    BAM file
"""

from argparse import ArgumentParser
from openpyxl import Workbook
from pysam import AlignmentFile
import sys

#sys.path.insert(0, '/bin/gt-seq/new_pipeline/')
#from gff3_manager import gff3_file

def argument_parser():
    parser = ArgumentParser(description='Options for soft-clipping filter.')
    parser.add_argument(
        '-b', '--bam',
        help="the bam file containing potentially soft-clipping reads. [stdin]",
        default = '-'
    )
    parser.add_argument(
        '-l', '--log',
        help="the log file to write the amount of reads."
    )
    #parser.add_argument(
    #    '-g', '--gff3',
    #    help="gff3 file with locations, to split counts by locus",
    #)
    parser.add_argument(
        '-o', '--output',
        help="the file path to write the filtered bam to [stdout]",
        default = '-'
    )
    return parser.parse_args()


class MappingStats:
    """
    class to hold information on the mapping quality.
    """
    def __init__(self, output_file):
        # TODO, count groups for each separate locus
        self.info_dict = {}
        self.outfile = output_file
        self.columns = ["read_group","total_reads","unmapped","multi_map","clipping","misbar","passing"]
        #self.gff3 = None
        self.loci = []

    def _grab_read_group_tag(self, segment):
        #TODO, getting read_group_tag this way only works becuase the RG_tagger script puts it on the end.
        return segment.tags[-1][1]

    def _grab_locus_tag(self, segment):
        #TODO: creating the locus id, might be better to cash. do with gff3_maniger?
        if segment.is_unmapped:
            return "other"
        else:
            locus = "{}_{:09}-{:09}".format(
                segment.reference_name,
                segment.reference_start,
                segment.reference_start+segment.reference_length
            )
            return locus

    #def add_gff3(self, gff3_filepath):
    #"""
    #can be used to get the primer set for a position
    #self.gff3.set_containing_position(segment.reference_name,segment.reference_start)
    #"""
    #    self.gff3 = gff3_file(gff3_filepath)

    def count(self, segment, column="total_reads"):
        """
        can be called for all the reads to get the total
        :param segment:
        """
        read_group_tag = self._grab_read_group_tag(segment)
        locus = self._grab_locus_tag(segment)
        try:
            self.info_dict[read_group_tag][column][locus] += 1
        except KeyError:
            if not read_group_tag in self.info_dict:
                self.info_dict[read_group_tag] = {}
            if not column in self.info_dict[read_group_tag]:
                self.info_dict[read_group_tag][column] = {}
            if not locus in self.info_dict[read_group_tag][column]:
                self.info_dict[read_group_tag][column][locus] = 1
                if not locus == 'other' and not locus in self.loci:
                    self.loci.append(locus)

    def clipping(self, segment):
        """
        can be called to count the amount of soft_clipping reads
        :param segment:
        """
        self.count(segment,column="clipping")

    def misbar(self, segment):
        self.count(segment,column="misbar")

    def multi(self, segment):
        self.count(segment,column="multi_map")

    def unmap(self, segment):
        self.count(segment,column="unmapped")

    def passing(self, segment):
        # the reads that have no problem, (a read can have multiple problems, so can't just subtract from total)
        self.count(segment,column="passing")

    def write(self):
        """
        figuring out in what way the data should be written to the outfile
        :return:
        """
        out_extention = self.outfile.split('.')[-1]
        if out_extention == 'txt' or out_extention == 'tsv':
            self.write_txt()
        elif out_extention == 'xlsx':
            self.write_xlsx()
        else:
            #TODO: as this is on the end it would suck extremely if nothing was written
            print("writing log as tab-seperated file")
            self.write_txt()

    def write_txt(self):
        """
        write the gathered counts to the log outfile
        """
        with open(self.outfile, 'w') as out:
            out.write('\t'.join(self.columns)+'\n')
            keys = sorted(self.info_dict.keys())
            for key in keys:
                value = self.info_dict[key]
                out.write(key + '\t')
                for x in self.columns[1:]:
                    if x in value:
                        out.write(str(sum(value[x].values()))+'\t')
                    else:
                        out.write(str(0)+'\t')
                out.write('\n')

    def write_xlsx(self):
        """
        write the output for each locus separate. using the self.outfile as a excel file
        :return:
        """
        wb = Workbook()
        rg_labels = sorted(self.info_dict.keys())
        loci_labels = []
        for locus in sorted(self.loci):
            loc_count = sum([self.info_dict[rg]['total_reads'][locus] for rg in rg_labels if locus in self.info_dict[rg]['total_reads']])
            if loc_count > 100:
                loci_labels.append(locus)
        # set sheet names
        sheet = wb.active
        sheet.title = self.columns[1]
        for sheet_name in self.columns[2:]:
            wb.create_sheet(sheet_name)
        # writing the data
        for sheet_name in self.columns[1:]:
            sheet = wb[sheet_name]
            # header
            _ = sheet.cell(column=2, row=1, value='other')
            for col,locus_id in enumerate(loci_labels):
                _ = sheet.cell(column=col + 3, row=1, value=locus_id)
            # data
            for row, rg_id in enumerate(rg_labels):
                if sheet_name in self.info_dict[rg_id]:
                    row_data = self.info_dict[rg_id][sheet_name]
                else:
                    row_data = {}
                _ = sheet.cell(column=1, row=row + 2, value=rg_id)
                other_count = sum([value for key,value in row_data.items() if not key in loci_labels])
                _ = sheet.cell(column=2, row=row + 2, value="{}".format(other_count))
                for col, locus_id in enumerate(loci_labels):
                    #if the locus_id is not there, it is not counted. and can be set to 0
                    if locus_id in row_data:
                        waarde = row_data[locus_id]
                    else:
                        waarde = 0
                    _ = sheet.cell(column=col + 3, row=row + 2, value="{}".format(waarde))
        wb.save(self.outfile)

def check_clipping(alignment,log=None):
    """
    checks if there are soft clipping bases in the segment.
    If the read is clipping, the read is filtered and returns False.

    :param alignment: pysam AlignedSegment
    :return: Boolean
    """
    if not alignment.is_unmapped:
        cigar = alignment.cigarstring
        if 'S' in cigar or 'H' in cigar:
            alignment.is_qcfail = True
            if log:
                log.clipping(alignment)
            return False
        else:
            return True
    else:
        return True


def check_barcode_is_off(alignment, tags, log=None):
    """
    See if the barcode was recognised with soft clipping.
    if so, it returns True and can be counted in the optional log

    :param alignment: the read
    :param tags: alignment tags as dict
    :return:
    """
    if 'RG' in tags:
        if tags['bm'] != '0':
            if log:
                log.misbar(alignment)
            return True
        else:
            return False
    else:
        return False


def check_multi_location(alignment, tags, log=None):
    """
    See if the read was mapped at multiple locations.
    if so, it returns True and can be counted in the optional log

    :param alignment: the read
    :param tags: alignment tags as dict
    :return:
    """
    if 'XA' in tags:
        alignment.is_qcfail = True
        if log:
            log.multi(alignment)
        return True
    else:
        return False


def check_is_mapped(alignment, log=None):
    if alignment.is_unmapped:
        if log:
            log.unmap(alignment)
        return True
    else:
        return False


def main(args):
    if args.log:
        log = MappingStats(args.log)
        #if args.gff3:
        #    log.add_gff3(args.gff3)
    else:
        log = None
    infile = AlignmentFile(args.bam, "rb")
    outfile = AlignmentFile(args.output, "wb", template=infile)
    for alignment in infile:
        tags = {x[0]: x[1] for x in alignment.tags}
        # checking and counting special cases
        c1 = check_multi_location(alignment, tags, log)
        c2 = check_clipping(alignment, log)
        if log:
            log.count(alignment)
            check_barcode_is_off(alignment, tags, log)
            c3 = check_is_mapped(alignment, log)
            if c1 or c2 or c3:
                # if any of these checks fail (return true), the read will not be counted in mpileup
                # if they all pass count it as passing
                log.passing(alignment)
        # writing the filtered sam file
        outfile.write(alignment)
    if args.log:
        log.write()


if __name__ == '__main__':
    args = argument_parser()
    main(args)