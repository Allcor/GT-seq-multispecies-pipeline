#!/usr/bin/python3
"""
Some samples are replicates of the same species. for a profile of the species replicates can be merged.
Figuring out those to merge is taken from the ReadGroup xlsm file.
merging is done with combining the read counts for on-target and allele.
The allele call should not be different but it is good to check.

Input:
  vcf file
  readgroup_sheet

Output:
  vcf file
"""

from argparse import ArgumentParser
from openpyxl import load_workbook
from collections import defaultdict
import logging
import vcf
import datetime
import sys

def argument_parser():
    parser = ArgumentParser(description='Options for vcf sample merger')
    parser.add_argument(
        '-v', '--vcf',
        required=True,
        help="the .vcf file with allele calls for all the read groups"
    )
    parser.add_argument(
        '-i', '--info',
        required=True,
        help="the .csv or .xlsx file with information on the read groups"
    )
    parser.add_argument(
        '-o', '--output',
        help="the path to the output file, will use sys.stdout if not given."
    )
    parser.add_argument(
        '-l', '--log',
        help="the log file to write the errors/comments",
        default="merge_samples.log"
    )
    parser.add_argument(
        '-s', '--species',
        help="only merge samples of given species"
    )
    return parser.parse_args()


def obtain_barcode_info_xlsx(info_file, species=None):
    info_lines = []
    wb = load_workbook(info_file, data_only=True)
    if 'Library' in wb.sheetnames:
        info_sheet = wb['Library']
    else:
        info_sheet = wb.active
    info_headers = []
    for column in info_sheet.iter_cols(max_row=1):
        info_headers.append(column[0].value)
    for row in info_sheet.iter_rows(min_row=2):
        info_dict = {info_headers[i]:item.value for i,item in enumerate(row)}
        # rows with empty cells are not allowed
        if all(x for x in info_dict.values()):
            if not species or info_dict['Species'].lower() == species.lower():
                info_lines.append(info_dict)
    return info_lines


def genotype_from_likelyhood_index(p,n,index):
    """
    figuring out the allele number corresponding to likelihood position
    ploidy P and N alternate alleles
    https://samtools.github.io/hts-specs/VCFv4.3.pdf
    :param p: the ploidy as int
    :param n: alt alleles as int
    :param index:
    :return: list with genotype numbers as strings (it's how pyvcf has it)
    """
    def recursive_order(_p, _n, alleles, suffix = []):
        for a in range(_n):
            if _p == 1:
                alleles.append([str(a)]+suffix)
            elif _p > 1:
                recursive_order(_p-1, a+1, alleles, [str(a)]+suffix)

    alleles_list = []
    recursive_order(p,n,alleles_list)
    return alleles_list[index]


def combine_samples(sample_list,new_name):
    """
    samples in the given list need to be merged and only one sample given back
    TODO: right now it changes one of the existing Call objects, this could be unwanted behavior.
    :param sample_list: list with vcf.model._Call objects
    :return: vcf.model._Call
    """
    merged_call = sample_list[0]
    for call in sample_list[1:]:
        #if the sample for this locus was not called it can be ignored.
        if call.called:
            # if the first one was none, take the called sample instead.
            if not merged_call.called:
                merged_call = call
            #combinging call data
            changes = {}
            # combine PL
            # "List of Phred-scaled genotype likelihoods", log10? and represent the likelihood of each allele (00,01,11).
            # using the largest score, it's the best I can find on biostars.
            # here is Koen of all people on how the PL is calculated. https://www.biostars.org/p/313799/
            if 'PL' in call.data._fields:
                old_PL = getattr(call.data, 'PL')
                new_PL = [max(old_PL[i],item) for i,item in enumerate(getattr(merged_call.data, 'PL'))]
                changes['PL'] = new_PL
            # the called genotype should be the same for merging
            # interpreting pred likelyhoods for the genotype otherwise
            if call.gt_bases != merged_call.gt_bases:
                min_index = new_PL.index(min(new_PL))
                pl_allele = genotype_from_likelyhood_index(call.ploidity,len(call.site.alleles),min_index)
                merged_call.gt_nums = '/'.join(pl_allele)
                merged_call.gt_alleles = pl_allele
            # sum DP
            if 'DP' in call.data._fields:
                new_DP = getattr(merged_call.data, 'DP') + getattr(call.data, 'DP')
                changes['DP'] = new_DP
            # combine SP,
            # "Phred-scaled strand bias P-value"
            # using the largest score, it's the best i can think of.
            if 'SP' in call.data._fields:
                new_SP = max(getattr(merged_call.data, 'SP'), getattr(call.data, 'SP'))
                changes['SP'] = new_SP
            # sum ADF
            if 'ADF' in call.data._fields:
                old_ADF = getattr(call.data, 'ADF')
                if isinstance(old_ADF,int):
                    new_ADF = getattr(merged_call.data, 'ADF') + old_ADF
                else:
                    new_ADF = [old_ADF[i]+item for i,item in enumerate(getattr(merged_call.data, 'ADF'))]
                changes['ADF'] = new_ADF
            # sum ADR
            if 'ADR' in call.data._fields:
                old_ADR = getattr(call.data, 'ADR')
                if isinstance(old_ADR,int):
                    new_ADR = getattr(merged_call.data, 'ADR') + old_ADR
                else:
                    new_ADR = [old_ADR[i]+item for i,item in enumerate(getattr(merged_call.data, 'ADR'))]
                changes['ADR'] = new_ADR
            # sum AD
            if 'AD' in call.data._fields:
                old_AD = getattr(call.data, 'AD')
                if isinstance(old_AD,int):
                    new_AD = getattr(merged_call.data, 'AD') + old_AD
                else:
                    new_AD = [old_AD[i]+item for i,item in enumerate(getattr(merged_call.data, 'AD'))]
                changes['AD'] = new_AD
            #replace the call data
            merged_call.data = merged_call.data._replace(**changes)
    merged_call.sample = new_name
    return merged_call


def merge_record(record,samples_to_merge):
    """
    merge the calls of the given record, samples to merge gives the calls to merge.
    :param record: vcf.model._Record
    :param samples_to_merge: dictionary with lists
    :return:
    """
    merged_record = record
    merged_samples = []
    #merge samples
    for key,sample_set in samples_to_merge.items():
        sample_index = [record._sample_indexes[str(s)] for s in sample_set]
        combined_sample = combine_samples([record.samples[i] for i in sample_index],key)
        merged_samples.append(combined_sample)
    merged_record.samples = merged_samples
    return merged_record


def main(args):
    logging.basicConfig(filename=args.log,level=logging.DEBUG)
    # open vcf input file
    vcf_in = vcf.Reader(filename=args.vcf)
    # add provenance
    vcf_in.metadata['MergeSamples'] = ['Merge samples from the same species']
    date = datetime.datetime.now().strftime("%a %b %d %X %Y")
    vcf_in.metadata['MergeCommand'] = ['{}; Date={}'.format(' '.join(sys.argv),date)]

    #open read group file and get samples to merge
    samples_to_merge = defaultdict(list)
    rg_info = obtain_barcode_info_xlsx(args.info, args.species)
    for sample in rg_info:
        samples_to_merge[sample['Cultivar_name']].append(sample['Sample_name'])

    #merge samples
    merged_records = []
    for record in vcf_in:
        merged_record = merge_record(record,samples_to_merge)
        merged_records.append(merged_record)

    # fix sample header
    #TODO: it's using the names with spaces and accented letters. Might give trouble.
    new_samples = list(samples_to_merge.keys())
    vcf_in.samples = new_samples

    # open vcf output file
    if args.output:
        outfile = open(args.output, 'w')
    else:
        outfile = sys.stdout
    vcf_out = vcf.Writer(outfile, vcf_in)
    for merged_record in merged_records:
        vcf_out.write_record(merged_record)

if __name__ == '__main__':
    args = argument_parser()
    main(args)