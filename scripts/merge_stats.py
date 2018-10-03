#!/usr/bin/python3
"""
When a library not only has replicates from a species but also duplicates from isolations the effectiveness
of merging can be tested.
This script takes the library information file and merges the duplicates / replicates.
Hopefully this will tell where this merging strategy needs improvement.
"""


from openpyxl import load_workbook, Workbook
from argparse import ArgumentParser
import vcf
import sys

sys.path.insert(0, '/bin/gt-seq/')
from new_pipeline import merge_samples


def argument_parser():
    parser = ArgumentParser(description='Options for vcf sample merger')
    parser.add_argument(
        '-v', '--vcf',
        required=True,
        help="the .vcf file with allele calls for all the read groups"
    )
    parser.add_argument(
        '-i', '--info',
        required=True,
        help="the .xlsx file with information on the read groups"
    )
    parser.add_argument(
        '-s', '--species',
        help="only merge samples of given species"
    )
    parser.add_argument(
        '-o', '--outfile',
        required=True,
        help="The path to write the new excel file to."
    )
    return parser.parse_args()


def obtain_barcode_info_xlsx(info_file, species=None):
    info_lines = []
    wb = load_workbook(info_file, data_only=True)
    if 'Library' in wb.sheetnames:
        info_sheet = wb['Library']
    else:
        info_sheet = wb.active
    info_headers = []
    for column in info_sheet.iter_cols(max_row=1):
        info_headers.append(column[0].value)
    for row in info_sheet.iter_rows(min_row=2):
        info_dict = {info_headers[i]:item.value for i,item in enumerate(row)}
        # rows with empty cells are not allowed
        if all(x for x in info_dict.values()):
            if not species or info_dict['Species'].lower() == species.lower():
                info_lines.append(info_dict)
    return info_lines


def create_pool_name(sample_name,box_name,position):
    #####
    #   Adjusting this will change how the samples are merged.
    #####
    pool_name = box_name
    return pool_name


def make_pools(sample_dict):
    """
    read the box numbers and cultivar names and create pools
    :param sample_dict:
    :return pools_list: list with Sample_ID to be merged
    """
    pools_dict = {}
    cultivar_list = []
    # first group them by cultivar
    for sample in sample_dict:
        cultivar = sample['Cultivar_name'].strip()
        if cultivar not in cultivar_list:
            cultivar_list.append(cultivar)
        if cultivar in pools_dict:
            pools_dict[cultivar].append(sample)
        else:
            pools_dict[cultivar] = [sample]
    # now group the duplicates from the cultivars together
    cultivar_list.sort()
    pools_returned = []
    for cultivar in cultivar_list:
        pool_sets = {}
        pool = pools_dict[cultivar]
        for sample in pool:
            pool_name = create_pool_name(sample['Cultivar_name'], sample['Plate_name'], sample['Position'])
            if pool_name in pool_sets:
                pool_sets[pool_name].append(sample['Sample_name'])
            else:
                pool_sets[pool_name] = [sample['Sample_name']]
            pool_sets[sample['Sample_name']] = sample
        pools_returned.append((cultivar,pool_sets))
    return pools_returned


def alleles_for_pools(vcf_object,pools_list):
    snp_dict = {}
    snp_list = []
    cultivar_list = set()
    for record in vcf_object:
        record_id = record.CHROM + '_{:09}'.format(record.POS)
        snp_list.append(record_id)
        cultivar_dict = {}
        for cultivar,pool in pools_list:
            seperate_values = {}
            to_merge = {}
            plate_order = []
            for key,read_group in pool.items():
                if isinstance(read_group, list):
                    # if it is a list these records should be merged
                    to_merge[key] = read_group
                else:
                    # it's a sample with the dict with info
                    call = record.genotype(key)
                    position = read_group['Plate_name']+'_'+ read_group['Position']
                    if read_group['Plate_name'] in seperate_values:
                        seperate_values[read_group['Plate_name']].append((position,call.gt_bases))
                    else:
                        seperate_values[read_group['Plate_name']] = [(position,call.gt_bases)]
                        plate_order.append(read_group['Plate_name'])
            # make sure there are two sets and they have less then 4 samples each, merge those.
            row = []
            for key,value in seperate_values.items():
                if len(value) > 4 or len(value) < 2:
                    dump = plate_order.pop(plate_order.index(key))
            if len(plate_order) < 2:
                #this sample is not added, hopefully for all snp
                pass
            else:
                plate_order.sort()
                if len(plate_order) > 2:
                    # if there are more then 2 plates, select 2 of them
                    plate_order = plate_order[:2]
                # add the normal reads
                for plate in plate_order:
                    samples_list = seperate_values[plate][:]
                    samples_list.sort()
                    for i in range(4):
                        if i < len(samples_list):
                            row.append(samples_list[i])
                        else:
                            row.append(('',''))
                for merge_name in plate_order:
                    groep = to_merge[merge_name]
                    calls = [record.genotype(name) for name in groep]
                    merged = merge_samples.combine_samples(calls,merge_name)
                    row.append((merge_name, merged.gt_bases))
                cultivar_list.add(cultivar)
                cultivar_dict[cultivar] = row
        snp_dict[record_id] = cultivar_dict
    cultivar_list = sorted(list(cultivar_list))
    return snp_dict,snp_list,cultivar_list


def main(args):
    sample_dict = obtain_barcode_info_xlsx(args.info,args.species)
    vcf_in = vcf.Reader(filename=args.vcf)
    # let's make the pools
    pools = make_pools(sample_dict)
    # get the (merged) allele calls
    snp_dict,snp_id,cultivar_id = alleles_for_pools(vcf_in,pools)
    # and creating a new excel for the output
    wb_out = Workbook()
    sheet = wb_out.active
    for x,snp in enumerate(snp_id):
        snp_offset = len(cultivar_id)*x
        for i,cultivar in enumerate(cultivar_id):
            line = snp_dict[snp][cultivar]
            _ = sheet.cell(row=i+1+snp_offset,column=1,value=snp)
            _ = sheet.cell(row=i+1+snp_offset,column=2,value=cultivar)
            for j,val in enumerate(line):
                _ = sheet.cell(row=i+1+snp_offset,column=j+3,value=val[1])
    wb_out.save(args.outfile)


if __name__ == '__main__':
    args = argument_parser()
    main(args)