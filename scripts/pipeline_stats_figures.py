#!/usr/local/bin/python3.6
"""
loads the pipeline stats excel sheet made by new_pipeline/pipeline_stats_xlsx.py
and creates plots with the data gathered in this file.
"""

from openpyxl import load_workbook
from argparse import ArgumentParser
import plotly.graph_objs as go
import plotly.offline as pyoff
import numpy
import scipy
from scipy import optimize, stats


def argument_parse():
    parser = ArgumentParser(description='Options for BarcodeSplit script')
    parser.add_argument(
        '-i', '--infile',
        required=True,
        help="the .xlsx file with statistics produced by the pipeline"
    )
    parser.add_argument(
        '-p', '--plots',
        help="The path of the folder to write the plots to."
    )
    return parser.parse_args()


html_template = """
<html>
<!-- Arlo Hoogeveen, Naktuinbouw 2018 -->
<head>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.1/css/bootstrap.min.css">
    <style>body{{ margin:0 100; background:whitesmoke; }}</style>
</head>
<body>
  <!-- Output from Plotly Python script -->
  <h1>gt_seq replicate samples data.</h1>
  {plot}
</body>
</html>
"""


def make_readdepth_regression_plot(loci_dict,average=True):
    # this requires the R^2 added to loci_dict in the 'make_comparison_plots' function
    layout = go.Layout(
        title='distribution of the regression fits',
        width=1500,
        height=640,
        hovermode='closest',
        legend=dict(
            orientation='h'
        ),
        yaxis=dict(
            title='R^2 value'
        ),
        xaxis=dict(
            title='total reads of '
        )
    )
    r_list = []
    c_list = []
    t_list = []
    for sample,value in loci_dict.items():
        info = value['info']
        if 'r_squared' in info:
            if average:
                r_list.append(sum(info['r_squared'])/len(info['r_squared']))
            else:
                r_list.append(max(info['r_squared']))
            c_list.append(int(info['count']))
            t_list.append(sample)
    #scatter trace
    scatter = go.Scatter(
        x = c_list,
        y = r_list,
        mode='markers',
        text=t_list,
        marker=dict(
            size=5,
            line=dict(
                width=0.5,
            )
        )
    )
    # fitting logaritmic function
    """
    #creating subset
    c_sub_list = []
    r_sub_list = []
    c_treshold = 300000
    for i,count in enumerate(c_list):
        if count < c_treshold:
            c_sub_list.append(count)
            r_sub_list.append(r_list[i].round(2))
    """
    # the math function as python function
    def fivepl(x, b, c):
        a = 0
        d = 1
        g = 0.25
        # https: // en.wikipedia.org / wiki / Generalised_logistic_function
        # https://stats.stackexchange.com/questions/190107/curve-fit-with-logarithmic-regression-in-python/338472#338472
        return (((a - d) / numpy.power(1 + (numpy.power((x / c),b)), g)) + d)
    #popt,pcov = scipy.optimize.curve_fit(lambda t,a,b: a+b*numpy.log(t), c_list, r_list)
    popt,pcov = scipy.optimize.curve_fit(fivepl, c_list, r_list, bounds=([0.1,1000],[5,10000]))
    fit_max = max(c_list)
    xi = numpy.arange(1, fit_max, (fit_max/100))
    line = fivepl(xi,*popt)
    fit_trace = go.Scatter(
        x=xi,
        y=line,
        mode='lines',
        marker=go.Marker(color='rgb(31, 119, 180)'),
        name='Fit'
    )
    layout_text = "growth rate = {:.2f}<br>inflection point = {:0f}".format(*popt)
    annotation1 = go.Annotation(
        x=0,
        y=1,
        xref='paper',
        yref='paper',
        yanchor='middle',
        xanchor='left',
        text=layout_text,
        align='left',
        font=go.Font(size=8),
        showarrow=False
    )
    #get the count cutoff
    cutoff = optimize.fsolve(lambda x: fivepl(x, *popt) - 0.965,1000)
    annotation2 = go.Annotation(
        x=cutoff[0],
        y=0.5,
        yanchor="middle",
        xanchor='left',
        text=str(int(cutoff[0])),
        align='left',
        showarrow=False
    )
    layout.update(
        shapes=[{
            'type': 'line',
            'x0': cutoff[0],
            'x1': cutoff[0]+1,
            'y0':0,
            'y1':1,
            'line': {
                'color': 'rgb(55, 128, 191)',
                'width': 3,
                'dash': 'dashdot',
            },
        }],
        annotations = [annotation1,annotation2]
    )
    # return the graph
    fig = dict(data=[scatter,fit_trace], layout=layout)
    div = pyoff.plot(fig, include_plotlyjs=False, output_type='div')
    return div


def read_stats_xlsx(info_file):
    wb = load_workbook(info_file, data_only=True)
    sheet = wb['mapping_x_sample']
    loci = ['other'] + [cell.value for cell in sheet[1]][2:]
    count_dict = {locus:{} for locus in loci}
    for row in sheet.iter_rows(min_row=2, min_col=1):
        sample = row[0].value
        for col,cell in enumerate(row[1:]):
            cell_dict = {key.strip("'"):int(value) for key,value in [x.split(':') for x in cell.value.strip('{}').split(',')]}
            count_dict[loci[col-1]][sample] = cell_dict
    return count_dict


def main(args):
    loci_dict = read_stats_xlsx(args.infile)
    r_div2 = make_readdepth_regression_plot(loci_dict, False)
    filename = args.plots + 'sample_{}.html'
    with open(filename.format('summery'), 'w') as outfile:
        outfile.write(html_template.format(plot=r_div2))


if __name__ == '__main__':
    args = argument_parse()
    main(args)