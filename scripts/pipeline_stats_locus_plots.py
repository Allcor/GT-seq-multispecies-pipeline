#!/usr/local/bin/python3.6
"""
loads the pipeline stats excel sheet made by new_pipeline/pipeline_stats_xlsx.py
and creates plots with the data gathered in this file.
"""

from openpyxl import load_workbook
from argparse import ArgumentParser
import plotly.graph_objs as go
from plotly import tools
import plotly.offline as pyoff
import numpy
from operator import itemgetter


def argument_parse():
    parser = ArgumentParser(description='Options for BarcodeSplit script')
    parser.add_argument(
        '-i', '--infile',
        required=True,
        help="the .xlsx file with statistics produced by the pipeline"
    )
    parser.add_argument(
        '-p', '--plots',
        required=True,
        help="The path of the folder to write the plots to."
    )
    parser.add_argument(
        '-s', '--separate',
        help="create separate files for each sequence variants instead of combining them for each locus."
    )
    parser.add_argument(
        '-u', '--url',
        help="The url to the genome on JBrowse (with the ?data=.. part)"
    )
    return parser.parse_args()


#####
# Traces
#####


def calculate_color(sample_values):
    # using a Triangular Number Sequence as offset
    # each time a different allel is added the offset is dependent by the
    # amount of combinations that can be made with alleles before
    #TODO: add some more colors
    COLORS = [
        "red",
        "purple",
        "blue",
        "orange",
        "green",
        "yellow",
        "black",
        "black",
        "black",
        "black",
        "black",
        "black"
    ]
    geno = [int(i) for i in sample_values['genotype'] if i.isdigit()]
    max_geno = max(geno) - 1
    offset = int((max_geno * (max_geno + 1)) / 2)
    color_nr = sum(geno) + offset
    return COLORS[color_nr]


def make_allele_depth_trace(samples_dict,jbrouwse_url=None):
    x = []
    y = []
    text = []
    color = []
    customdata = []
    for sample,values in samples_dict.items():
        x.append(values['x'])
        y.append(values['y'])
        text.append(sample)
        # color accounting for multiple alt alleles
        if values['genotype'] != 'None':
            value_color = calculate_color(values)
        else:
            value_color = 'gray'
        color.append(value_color)
        if jbrouwse_url:
            customdata.append(jbrouwse_url.give_url(track=values['read_group_id']))
    depth_trace = go.Scatter(
        x=x,
        y=y,
        mode='markers',
        customdata=customdata,
        text=text,
        marker=dict(
            size=5,
            color=color,
            line=dict(
                width=0.5,
            )
        )
    )
    return depth_trace


def make_allele_percent_trace(samples_dict,jbrouwse_url=None):
    x = []
    y = []
    text = []
    color = []
    customdata = []
    for sample,values in samples_dict.items():
        x.append(values['y']+values['x'])
        y.append(values['p_A2'])
        text.append(sample)
        # color accounting for multiple alt alleles
        if values['genotype'] != 'None':
            value_color = calculate_color(values)
        else:
            value_color = 'gray'
        color.append(value_color)
        if jbrouwse_url:
            customdata.append(jbrouwse_url.give_url(track=values['read_group_id']))
    depth_trace = go.Scatter(
        x=x,
        y=y,
        mode='markers',
        customdata=customdata,
        text=text,
        marker=dict(
            size=5,
            color=color,
            line=dict(
                width=0.5,
            )
        )
    )
    return depth_trace


def make_mapping_on_target_trace(read_count_dict, size_order=False, loci_to_emphasize=()):
    x = []
    y = []
    text = []
    color = []
    st_dev = []
    for locus,samples_dict in read_count_dict.items():
        total = sum([data['total_reads'] for data in samples_dict.values()])
        usefull = sum([data['passing'] for data in samples_dict.values()])
        individual_fractions = []
        if locus in loci_to_emphasize:
            color.append('red')
        else:
            color.append('#1f77b4')  # muted blue
        for data in samples_dict.values():
            if data['total_reads'] != 0:
                individual_fractions.append(data['passing'] / data['total_reads'])
            else:
                individual_fractions.append(0)
        st_dev.append(numpy.std(individual_fractions))
        x.append(locus)
        y.append(usefull/total)
        text.append(locus+"<br>{}/{}".format(usefull,total))
    # if the table has to be sorted by size,
    if size_order:
        order = [i for i,val in sorted(enumerate(y), key=lambda pair: pair[1])]
        x = [x[i] for i in order]
        y = [y[i] for i in order]
        text = [text[i] for i in order]
        color = [color[i] for i in order]
        st_dev = [st_dev[i] for i in order]
    # and creating the plot itself
    on_target_trace = go.Bar(
        x=x,
        y=y,
        text=text,
        error_y = dict(
            type='data',
            array=st_dev,
            thickness=1,
            color='rgba(black,0.5)',
            visible=True
        ),
        marker=dict(
            color=color
        )
    )
    return on_target_trace


def make_read_depth_trace(read_count_dict, size_order=False, loci_to_emphasize=()):
    x = []
    y = []
    text = []
    color = []
    st_dev = []
    # let's gather the required values
    total_on_target = []
    for locus, samples_dict in read_count_dict.items():
        passing = [data['total_reads'] for data in samples_dict.values()]
        on_target_locus = sum(passing)
        #TODO: this is basically the standardeviation nate has, with the total and percentage things get strange.
        st_dev_locus = numpy.std([(x/on_target_locus) for x in passing])
        total_on_target.append((locus,on_target_locus,st_dev_locus))
    on_target_getter = itemgetter(1)
    total = sum([data for data in map(on_target_getter,total_on_target)])
    # sorting if required
    if size_order:
        total_on_target.sort(key=on_target_getter)
    # putting them in the lists required for the plot
    for locus,on_target_locus,st_dev_locus in total_on_target:
        x.append(locus)
        y.append(on_target_locus/total)
        text.append(locus + "<br>{}/{}".format(on_target_locus, total))
        if locus in loci_to_emphasize:
            color.append('red')
        else:
            color.append('#1f77b4')  # muted blue
        st_dev.append(st_dev_locus)
    # and creating the plot itself
    on_target_trace = go.Bar(
        x=x,
        y=y,
        text=text,
        error_y = dict(
            type='data',
            array=st_dev,
            thickness=1,
            color='rgba(black,0.5)',
            visible=True
        ),
        marker=dict(
            color=color
        )
    )
    return on_target_trace

#####
# making the figures
#####

def make_snp_plots(samples_dict,snp_name="",jbrouwse_url=None):
    height = 1 * 350
    width = 2 * 350
    # if there are multiple plots the space in between has to be acounted for.
    width += width*0.2
    #height += height*0.3
    # and the borders
    height += 180
    width += 160
    fig = tools.make_subplots(
        rows=1,
        cols=2,
        subplot_titles=('allele depth', 'allele 2 percent')
    )
    fig['layout'].update(
        title="sequence variant {}.".format(snp_name),
        width=width,
        height=height,
        hovermode='closest',
        showlegend=False,
        annotations=[
            dict(
                x=0.48,
                y=1,
                xref='paper',
                yref='paper',
                yanchor='middle',
                xanchor='right',
                text='lines do not reflect SNP calling rules',
                font=go.Font(size=8),
                showarrow=False,
            ),
            dict(
                x=1,
                y=1,
                xref='paper',
                yref='paper',
                yanchor='middle',
                xanchor='right',
                text='lines do not reflect SNP calling rules',
                font=go.Font(size=8),
                showarrow=False,
            )
        ],
        shapes = [
            {
                'type': 'line',
                'x0': 0,
                'x1': 10,
                'xref' : "x",
                'y0': 10,
                'y1': 0,
                'yref' : "y",
                'line': {
                    'color': 'yellow',
                    'width': 2,
                },
            },{
                'type': 'line',
                'x0': 9,
                'x1': 10000,
                'xref': "x",
                'y0': 1,
                'y1': 1000,
                'yref': "y",
                'line': {
                    'color': 'red',
                    'width': 2,
                },
            },{
                'type': 'line',
                'x0': 1000,
                'x1': 1,
                'xref': "x",
                'y0': 10000,
                'y1': 9,
                'yref': "y",
                'line': {
                    'color': 'blue',
                    'width': 2,
                }
            },{
                'type': 'line',
                'x0': 5,
                'x1': 10000,
                'xref': "x",
                'y0': 5,
                'y1': 10000,
                'yref': "y",
                'line': {
                    'color': 'purple',
                    'width': 2,
                }
            },{
                'type': 'line',
                'x0': 6.6,
                'x1': 10000,
                'xref': "x",
                'y0': 3.3,
                'y1': 5000,
                'yref': "y",
                'line': {
                    'color': 'black',
                    'width': 2,
                }
            },{
                'type': 'line',
                'x0': 3.3,
                'x1': 5000,
                'xref': "x",
                'y0': 6.6,
                'y1': 10000,
                'yref': "y",
                'line': {
                    'color': 'black',
                    'width': 2,
                }
            },
            # and now the percentage plot
            {
                'type': 'line',
                'x0': 10,
                'x1': 10,
                'xref': "x2",
                'y0': 0,
                'y1': 100,
                'yref': "y2",
                'line': {
                    'color': 'yellow',
                    'width': 2,
                }
            },{
                'type': 'line',
                'x0': 0,
                'x1': 10000,
                'xref': "x2",
                'y0': 10,
                'y1': 10,
                'yref': "y2",
                'line': {
                    'color': 'red',
                    'width': 2,
                }
            },{
                'type': 'line',
                'x0': 0,
                'x1': 10000,
                'xref': "x2",
                'y0': 90,
                'y1': 90,
                'yref': "y2",
                'line': {
                    'color': 'blue',
                    'width': 2,
                }
            },{
                'type': 'line',
                'x0': 0,
                'x1': 10000,
                'xref': "x2",
                'y0': 66.6,
                'y1': 66.6,
                'yref': "y2",
                'line': {
                    'color': 'black',
                    'width': 2,
                }
            },{
                'type': 'line',
                'x0': 0,
                'x1': 10000,
                'xref': "x2",
                'y0': 33.3,
                'y1': 33.3,
                'yref': "y2",
                'line': {
                    'color': 'black',
                    'width': 2,
                }
            }
        ],
    )
    #make both axes the same scale
    max_depth = max([sum([v['x'],v['y']]) for v in samples_dict.values()])
    fig['layout']['xaxis1'].update(
        range=[max_depth*-0.05, max_depth*1.05],
        title='Reference depth (vcf)'
    )
    fig['layout']['yaxis1'].update(
        range=[max_depth*-0.05, max_depth*1.05],
        title='Alternate depth (vcf)'
    )
    fig['layout']['xaxis2'].update(
        range=[max_depth * -0.05, max_depth * 1.05],
        title='depth (vcf)'
    )
    fig['layout']['yaxis2'].update(
        tickvals=[0, 25, 50, 75, 100],
        ticktext=['0%', '', '50%', '', '100%'],
        title='percent alternate allele'
    )
    # making the traces themselves
    fig.append_trace(make_allele_depth_trace(samples_dict,jbrouwse_url=jbrouwse_url),1,1)
    fig.append_trace(make_allele_percent_trace(samples_dict,jbrouwse_url=jbrouwse_url),1,2)
    locus_div = pyoff.plot(fig, include_plotlyjs=False, output_type='div')
    return locus_div


def make_distribution_plots(read_count_dict, loci_to_emphasize=()):
    height = 1*250
    width = 2*350
    # if there are multiple plots the space in between has to be acounted for.
    width += width * 0.2
    # height += height*0.3
    # and the borders
    height += 180
    width += 160
    fig = tools.make_subplots(
        rows=1,
        cols=2,
        subplot_titles=('read_distribution_among_loci', 'mapping_on_target')
    )
    fig['layout'].update(
        width=width,
        height=height,
        showlegend=False,
    )
    fig['layout']['xaxis1'].update(
        title = "Loci",
        showticklabels = False
    )
    fig['layout']['yaxis1'].update(
        title='fraction of total on-target reads',
        rangemode="nonnegative"
    )
    fig['layout']['xaxis2'].update(
        title = "Loci",
        showticklabels = False
    )
    fig['layout']['yaxis2'].update(
        title='fraction on-target reads that are valid'
    )
    # making the plots themseleves
    fig.append_trace(make_read_depth_trace(
        read_count_dict,
        size_order=True,
        loci_to_emphasize=loci_to_emphasize
    ),1,1)
    fig.append_trace(make_mapping_on_target_trace(
        read_count_dict,
        size_order=True,
        loci_to_emphasize=loci_to_emphasize
    ),1,2)
    distribution_div = pyoff.plot(fig, include_plotlyjs=False, output_type='div')
    return distribution_div

#####
# Read input data
#####

def loci_stats_xlsx(excel_workbook):
    loci_dict = {}
    sheet = excel_workbook['loci_info']
    for row in sheet.iter_rows(min_row=2,min_col=1):
        locus_id = row[0].value
        snp_set = row[1].value.split(',')
        main_snp = row[2].value
        loci_dict[locus_id] = (main_snp,snp_set)
    return loci_dict


def read_stats_xlsx(excel_workbook):
    # the sheet containing the call stats
    sheet = excel_workbook['mapping_x_sample']
    loci = [cell.value for cell in sheet[1]][1:]
    # create a dictionary for each sample and locus combination containing the calls as in the vcf file
    reads_dict = {locus:{} for locus in loci}
    for row in sheet.iter_rows(min_row=2, min_col=1):
        sample = row[0].value
        for col,cell in enumerate(row[1:]):
            cell_list = [x.split(':') for x in cell.value.strip('{}').split(',')]
            cell_dict = {key.strip().strip("'"):value.strip() for key,value in cell_list}
            for key,value in cell_dict.items():
                if value.isdigit():
                    cell_dict[key] = int(value)
            reads_dict[loci[col]][sample] = cell_dict
    return reads_dict


def readgroup_xlsx(excel_workbook):
    readgroups_dict = {}
    sheet = excel_workbook['sample_info']
    header = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2,min_col=1):
        values_dict = {header[i]:cell.value for i,cell in enumerate(row)}
        readgroups_dict[values_dict['sample_name']] = values_dict['read_group_id']
    return readgroups_dict


def call_stats_xlsx(excel_workbook):
    # the sheet containing the call stats
    sheet = excel_workbook['snp_x_samples']
    snp = [cell.value for cell in sheet[1]][1:]
    # create a dictionary for each sample and locus combination containing the calls as in the vcf file
    call_dict = {locus:{} for locus in snp}
    for row in sheet.iter_rows(min_row=2, min_col=1):
        sample = row[0].value
        for col,cell in enumerate(row[1:]):
            cell_list = [x.split(':') for x in cell.value.strip('{}').split(',')]
            cell_dict = {key.strip().strip("'"):value.strip() for key,value in cell_list}
            for key,value in cell_dict.items():
                if value.isdigit():
                    cell_dict[key] = int(value)
            call_dict[snp[col]][sample] = cell_dict
    return call_dict

class JBrowseUrl:
    def __init__(self,base_url=None):
        self.base = ""
        self.data = ""
        self.loc = ""
        self.tracks = []
        self.highlight = ""
        if base_url:
            self.set_params_from_url(base_url)

    def set_params_from_url(self,base_url):
        self.base = base_url.split('?')[0]
        get_stuff = base_url.split('?')[1].split('&')
        for get_part in get_stuff:
            split_part = get_part.split('=')
            # The path to the genome where all the tracks are saved
            if split_part[0] == 'data':
                self.data = split_part[1].replace('%2F','/')
            # Position on the genome
            if split_part[0] == 'loc':
                self.loc = split_part[1].replace('%3A',':')
            # Tracks to show
            if split_part[0] == 'tracks':
                self.tracks = split_part[1].replace('%2C',',').split(',')
            # Highlighted region
            if split_part[0] == 'highlight':
                self.highlight = split_part[1].replace('%3A',':')

    def set_location(self,chr,start,stop):
        self.loc = chr+':'+start+'..'+stop

    def give_url(self, **kwargs):
        # base html location
        if 'base' in kwargs:
            base = kwargs['base']
        else:
            base = self.base
        # the path to the genome
        if 'data' in kwargs:
            data = kwargs['data']
        else:
            data = self.data
        # the location
        if 'loc' in kwargs:
            loc = kwargs['loc']
        else:
            loc = self.loc
        # the tracks
        if 'tracks' in kwargs:
            tracks = kwargs['tracks']
        else:
            tracks = self.tracks[:]
        if 'track' in kwargs:
            tracks.append(kwargs['track'])
        tracks_str = ','.join(tracks)
        # the highlighted region
        if 'highlight' in kwargs:
            highlight = kwargs['highlight']
        else:
            highlight = self.highlight
        # and make the url
        url_template = "{base}?data={data}&loc={loc}&tracks={tracks}&highlight={highlight}"
        return url_template.format(base=base,data=data,loc=loc,tracks=tracks_str,highlight=highlight)

#####
# putting it all together
#####


def make_all_locus_plots(filename,loci_dict,read_dict,snp_dict,jbrouwse_url=None):
    html_template = """
    <html>
    <!-- Arlo Hoogeveen, Naktuinbouw 2018 -->
    <head>
        <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" integrity="sha384-Gn5384xqQ1aoWXA+058RXPxPg6fy4IWvTNh0E263XmFcJlSAwiGgFAW/dAiS6JXm" crossorigin="anonymous">
        <style>body{{ margin:0 100; background:whitesmoke; }}</style>
    </head>
    <body>
      <!-- Output from Plotly Python script -->
      {title}
      {plot}
      {distribution}

      <!-- Optional JavaScript -->
      <!-- jQuery first, then Popper.js, then Bootstrap JS -->
      <script src="https://code.jquery.com/jquery-3.2.1.slim.min.js" integrity="sha384-KJ3o2DKtIkvYIK3UENzmM7KCkRr/rE9/Qpg6aAZGJwFDMVNA/GpGFF93hXpG5KkN" crossorigin="anonymous"></script>
      <script src="https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" integrity="sha384-ApNbgh9B+Y1QKtv3Rn7W3mgPxhU9K/ScQsAP7hUibX39j7fakFPskvXusvfa0b4Q" crossorigin="anonymous"></script>
      <script src="https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" integrity="sha384-JZR6Spejh4U02d8jOt6vLEHfe/JQGiRRSQQxSfFWpi1MquVdAyjUar5+76PVCmYl" crossorigin="anonymous"></script>
      {script}
    </body>
    </html>
    """
    script = """
    <script>
        var plots = document.getElementsByClassName('plotly-graph-div');
        var i;
        for (i = 0; i < plots.length; i++){
            var plot = plots[i];
            plot.on('plotly_click', function(data){
                for(var i=0; i < data.points.length; i++){
                    if ('customdata' in data.points[i]){
                        var link = data.points[i].customdata
                        window.open(link);
                    }
                }
            });
        };
    </script>
    """
    locus_read_dict = {}
    for locus, samples_dict in read_dict.items():
        if locus in loci_dict:
            locus_read_dict[locus] = samples_dict
        # TODO: probably want to do stuff with other and rogue
    #making the plots for each locus
    for locus,info in loci_dict.items():
        #setting the loc of jbrouwse
        if jbrouwse_url:
            split_loc = locus.split('_')
            split_pos = split_loc[1].split('-')
            jbrouwse_url.set_location(split_loc[0],str(int(split_pos[0])-20),str(int(split_pos[1])+20))
        # first the plots of the snp calls among samples
        snp_id_list = info[1]
        main_snp_id = snp_id_list.pop(snp_id_list.index(info[0]))
        main_snp_div = make_snp_plots(snp_dict[main_snp_id],snp_name=main_snp_id,jbrouwse_url=jbrouwse_url)
        if snp_id_list != []:
            other_snp_div = []
            for snp_id in snp_id_list:
                other_snp_div.append(make_snp_plots(snp_dict[snp_id],snp_name=snp_id,jbrouwse_url=jbrouwse_url))
            snp_plots_html = """
            {main_snp}
            <p>
              <a class="btn btn-primary" data-toggle="collapse" href="#collapseExample" role="button" aria-expanded="false" aria-controls="collapseExample">
                More sequence variants
              </a>
            </p>
            <div class="collapse" id="collapseExample">
              {other_snp}
            </div>
            """
            snp_div = snp_plots_html.format(main_snp=main_snp_div,other_snp="<br>".join(other_snp_div))
        else:
            snp_div = main_snp_div
        # and the distribution plots
        distribution_div = make_distribution_plots(locus_read_dict, loci_to_emphasize=locus)
        # and writing the output file.
        with open(filename.format(locus), 'w') as outfile:
            outfile.write(html_template.format(
                plot=snp_div,
                distribution=distribution_div,
                title="<h1>genotyping results on locus {}</h1>".format(locus),
                script=script
            ))


def make_all_snp_plots(filename,read_dict,snp_dict):
    html_template = """
    <html>
    <!-- Arlo Hoogeveen, Naktuinbouw 2018 -->
    <head>
        <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" integrity="sha384-Gn5384xqQ1aoWXA+058RXPxPg6fy4IWvTNh0E263XmFcJlSAwiGgFAW/dAiS6JXm" crossorigin="anonymous">
        <style>body{{ margin:0 100; background:whitesmoke; }}</style>
    </head>
    <body>
      <!-- Output from Plotly Python script -->
      {title}
      {plot}
      {distribution}
    </body>
    </html>
    """
    loci_dict = {}
    for locus, samples_dict in read_dict.items():
        if 'other' in locus or 'rogue' in locus:
            # TODO: probably want to do stuff with other and rogue
            pass
        else:
            loci_dict[locus] = samples_dict
    # and make the plots
    for snp, samples_dict in snp_dict.items():
        locus_div = make_snp_plots(samples_dict)
        locus = [x['locus'] for x in samples_dict.values()][0]
        distribution_div = make_distribution_plots(loci_dict, loci_to_emphasize=locus)
        with open(filename.format(snp), 'w') as outfile:
            outfile.write(html_template.format(
                plot=locus_div,
                distribution=distribution_div,
                title="<h1>genotyping results on locus {}</h1>".format(snp)
            ))

def main(args):
    filename = args.plots + 'locus_{}.html'
    # open the exel file
    wb = load_workbook(args.infile, data_only=True)
    # getting the snp calls.
    snp_dict = call_stats_xlsx(wb)
    # getting the number of reads
    read_dict = read_stats_xlsx(wb)
    # getting the locus
    loci_dict = loci_stats_xlsx(wb)
    # checking if jbrowse links are possible
    if args.url:
        jbrowse_url = JBrowseUrl(args.url)
        # adding the readgroup id to the snp_dict to use as track lable
        readgroups_dict = readgroup_xlsx(wb)
        for snp_id,sample_dict in snp_dict.items():
            for sample_id,info in sample_dict.items():
                info['read_group_id'] = readgroups_dict[sample_id]
    else:
        jbrowse_url = None
    # making the plots
    if args.separate:
        # not default, would make a page for each snp
        make_all_snp_plots(filename,read_dict,snp_dict)
    else:
        make_all_locus_plots(filename,loci_dict,read_dict,snp_dict,jbrowse_url)


if __name__ == '__main__':
    args = argument_parse()
    main(args)