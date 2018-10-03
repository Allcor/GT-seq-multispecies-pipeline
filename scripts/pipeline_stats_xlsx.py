#!/usr/local/bin/python3.6
"""
read a vcf file and get the allel persentage for all loci and sample combinations
"""

from openpyxl import load_workbook, Workbook
from argparse import ArgumentParser
import scipy
from scipy import stats
import vcf
import sys

# TODO: set pithonpath with install
sys.path.insert(0, '/bin/gt-seq/')
from new_pipeline import gff3_manager

def argument_parse():
    parser = ArgumentParser(description='Options for BarcodeSplit script')
    parser.add_argument(
        "-v", "--vcf",
        dest="vcf_file",
        required=True,
        help="the path to the vcf file containing varients of the GT-seq run."
    )
    parser.add_argument(
        '-i', '--info',
        required=True,
        help="the .csv or .xlsx file with information on the read groups"
    )
    parser.add_argument(
        '-g', '--gff3',
        required=True,
        help="positional information on the primers and variants."
    )
    parser.add_argument(
        '-o', '--outfile',
        required=True,
        help="The path to write the new excel file to."
    )
    parser.add_argument(
        '-c', '--counts',
        required=True,
        help="the file containing the read counts"
    )
    return parser.parse_args()


def summeryFigures_allel_calculation(loci):
    sample_dict = {}
    for sample in loci.samples:
        values = {}
        # size = len(info)
        # if size == 13:
        #    A1_corr = float(info[6])
        #    A2_corr = float(info[7])
        # xarr = info[1].split('=')
        if isinstance(sample.data.AD, list):
            xarr = [loci.alleles[0], sample.data.AD[0]]
            try:
                yarr = [loci.alleles[1].sequence, sample.data.AD[1]]
            except AttributeError:
                yarr = [loci.alleles[1].type, sample.data.AD[1]]
        else:
            xarr = [loci.alleles[0], sample.data.AD]
            yarr = ['.', 0]

        # yarr = info[2].split('=')
        values['x'] = int(round(float(xarr[1])))
        values['y'] = int(round(float(yarr[1])))
        values['genotype'] = sample.gt_nums
        AF_div = float(xarr[1]) + float(yarr[1])
        if AF_div == 0:
            AF_div = 0.1
        values['p_A2'] = (float(yarr[1]) / AF_div) * 100
        sample_dict[sample.sample] = values
    return sample_dict


def obtain_barcode_info_xlsx(info_file):
    info_lines = []
    wb = load_workbook(info_file, data_only=True)
    if 'Library' in wb.sheetnames:
        info_sheet = wb['Library']
    else:
        info_sheet = wb.active
    info_headers = []
    for column in info_sheet.iter_cols(max_row=1):
        info_headers.append(column[0].value)
    for row in info_sheet.iter_rows(min_row=2):
        info_dict = {info_headers[i]:item.value for i,item in enumerate(row)}
        # rows with empty cells are not allowed
        if all(x for x in info_dict.values()):
            info_lines.append(info_dict)
    return info_lines


def combine_count_totals(counts_dict):
    read_counts = {sample: {} for sample in counts_dict.keys()}
    get_total = ['total_reads', 'unmapped', 'misbar']
    for sample_name in counts_dict.keys():
        for stats_label in get_total:
            count = sum([int(x) for x in counts_dict[sample_name][stats_label].values()])
            read_counts[sample_name][stats_label] = count
    return read_counts


def obtain_counts_xlsx(excel_file,sample_id_list):
    # for total_reads, unmapped and misbar only return total
    read_counts = {sample:{} for sample in sample_id_list}
    wb = load_workbook(excel_file, data_only=True)
    for stats in wb.sheetnames:
        sheet = wb[stats]
        header = [x.value for x in sheet[1]][1:]
        for row in sheet.iter_rows(min_row=2):
            sample_name = '_'.join(row[0].value.split('_')[2:])
            if sample_name in sample_id_list:
                count_dict = {header[i]:int(value.value) for i,value in enumerate(row[1:])}
                read_counts[sample_name][stats] = count_dict
    return read_counts


def combine_count_positions(counts_dict, gff3=None):
    counts_dict_combined = {}
    #get key's
    combine_header = set()
    combine_locus = set()
    for sample,value in counts_dict.items():
        for value_key in value.keys():
            combine_header.add(value_key)
            # TODO "total_reads" should always be availeble, still might be better to get first element from combine_header
            for locus_key in value["total_reads"].keys():
                if locus_key != "other":
                    combine_locus.add(locus_key)
    # get the location names, and counts to pick the best name
    sum_dict = {locus:0 for locus in combine_locus}
    for sample,value in counts_dict.items():
        # soft-clipping could mess up finding the 'proper' alignment
        for locus_key in combine_locus:
            total_count = value["total_reads"][locus_key]
            soft_clip_count = value["clipping"][locus_key]
            sum_dict[locus_key] += (total_count-soft_clip_count)
    # combine the counts of locations close to one another
    locus_keys = sorted(sum_dict.keys())
    locus_to_pool = {"other":["other"]}
    group_name = ""
    pre_chr = ""
    pre_start = -10
    pre_end = -10
    for locus_id in locus_keys:
        chr = '_'.join(locus_id.split('_')[:-1])
        start,stop = (int(pos) for pos in locus_id.split('_')[-1].split('-'))
        if pre_chr != chr:
            pre_chr = chr
            pre_start = -10
            pre_end = -10
        if pre_start-10 < start > pre_start+10 and pre_end-10 < stop > pre_end+10:
            pre_start = start
            pre_end = stop
            group_name = locus_id
            locus_to_pool[group_name] = []
        elif sum_dict[locus_id] > sum_dict[group_name]:
            locus_to_pool[locus_id] = locus_to_pool.pop(group_name)
            pre_start = start
            pre_end = stop
            group_name = locus_id
        locus_to_pool[group_name].append(locus_id)
    # check the locations in gff3
    found_amplicons = {}
    if gff3:
        target_amplicons = {}
        for feature_id, gff_feature in gff3.features_id.items():
            if gff_feature.type == "PCR_product" and gff_feature.attributes.active:
                target_amplicons[feature_id] = gff_feature
        for key, list in locus_to_pool.items():
            if not key == "other":
                #TODO: this seems to be a bit over complicated but because the most occuring sequence end and start
                #TODO: not necessarily needs to be the target pcr product i made some checks,
                #TODO: checks just are a bit indirect
                chr = '_'.join(key.split('_')[:-1])
                start,stop = (int(pos) for pos in key.split('_')[-1].split('-'))
                gff_set = gff3.set_containing_position(chr,start,stop)
                if gff_set and len(gff_set.pcr_product) != 0:
                    # remove the feature from target's so we can see how many are not found.
                    product_id = None
                    for product in gff_set.pcr_product:
                        # check if amplicons are not overlapping
                        if start-2 <= product.start <= start+2 and stop-2 <= product.end <= stop+2:
                            product_id = product.attributes.id
                    if not product_id:
                        # something strange going on, the right annotation was not found
                        found_amplicons[key] = 'rogue'
                        #product_id = gff_set.pcr_product[0].attributes.id
                    elif product_id not in target_amplicons:
                        found_amplicons[key] = product_id+'_filtered'
                    else:
                        # or a filtered read was still sequenced
                        found_amplicons[key] = target_amplicons.pop(product_id).attributes.id
                else:
                    found_amplicons[key] = "rogue"
        if target_amplicons:
            found_amplicons["other"] = ', '.join([foo.attributes.id for i,foo in target_amplicons.items()]+['and other'])
        else:
            found_amplicons["other"] = "other"
    # add the counts
    for location, group in locus_to_pool.items():
        if found_amplicons != {}:
            # add the original name as lable
            locus_label = found_amplicons[location]
            if locus_label == "rogue":
                locus_label += " {}".format(location)
        else:
            locus_label = location
        counts_dict_combined[locus_label] = {}
        for sample in counts_dict:
            sample_dict = {}
            for data in combine_header:
                sample_dict[data] = sum([int(counts_dict[sample][data][x]) for x in group])
            counts_dict_combined[locus_label][sample] = sample_dict
    return counts_dict_combined


def compare_sample_calls(loci_list,sample_id_list,loci_dict):
    compare_dict = {}
    for sample_x in sample_id_list:
        compare_dict[sample_x] = {}
        for sample_y in sample_id_list:
            x = []
            y = []
            similarity = []
            for locus in loci_list:
                site_x = loci_dict[str(sample_x)][locus]
                x.append(round(site_x['p_A2'], 1))
                site_y = loci_dict[str(sample_y)][locus]
                y.append(round(site_y['p_A2'], 1))
                # checking if it has the same call
                # TODO: Scoring all match the same, even if no call could be made.
                # TODO: Also not all mismatch is the same, A/A and A/T should be closer then A/A and T/T.
                if site_x['genotype'] == site_y['genotype']:
                    similarity.append(1)
                else:
                    similarity.append(0)
            # compare calls
            call_percent = sum(similarity) / len(similarity)
            # create the linear fit
            slope, intercept, r_value, p_value, std_err = scipy.stats.linregress(x, y)
            compare_dict[sample_x][sample_y] = {
                'slope':slope,
                'intercept':intercept,
                'r_value':r_value,
                'p_value':p_value,
                'std_err':std_err,
                'call_similarity':call_percent,
            }
    return compare_dict


def create_sheet_genotype_data(workbook,loci_dict,rows=None,columns=None):
    sheet = workbook.create_sheet('snp_x_samples')
    if not rows:
        rows = loci_dict.keys()
    if not columns:
        columns = loci_dict[rows[0]].keys()
    #header
    for col, locus in enumerate(columns):
        _ = sheet.cell(column=col+2, row=1, value=locus)
    #values
    for row,sample in enumerate(rows):
        _ = sheet.cell(column=1, row=row+2, value=sample)
        for col,locus in enumerate(columns):
            _ = sheet.cell(column=col+2, row=row+2, value="{}".format(loci_dict[sample][locus]))


def write_sheet_main(workbook, sample_id_list, rg_info, compare_dict, samples_to_merge, counts):
    sheet = workbook.active
    sheet.title = "sample_info"
    header = ['sample_name',
              'read_group_id',
              'merge_group',
              'box_positie',
              'read_count',
              'not_mapped_reads',
              'fuzzy_barcodes',
              'closest_neighbour']
    _rg_info = {sample['Sample_name']:sample for sample in rg_info}
    for i,lable in enumerate(header):
        sheet.cell(column=i+1, row=1,value=lable)
    for row,sample_name in enumerate(sample_id_list):
        sample = _rg_info[sample_name]
        # sample_name
        sheet.cell(column=1, row=row+2, value=sample_name)
        # readgroup
        read_group_id = sample['Sample_ID']
        sheet.cell(column=2, row=row+2, value=read_group_id)
        # merge_group
        merge_group = samples_to_merge[sample['Cultivar_name']]
        sheet.cell(column=3, row=row + 2, value='{}'.format(merge_group))
        # box_positie
        box_positie = sample['Plate_name'] + '_' + sample['Position']
        sheet.cell(column=4, row=row+2, value=box_positie)
        # read_count
        count_info = counts[sample_name]
        sheet.cell(column=5, row=row+2, value=count_info['total_reads'])
        # unmapped
        sheet.cell(column=6, row=row + 2, value=count_info['unmapped'])
        # fuzzy_barcodes
        if 'misbar' in count_info:
            sheet.cell(column=7, row=row+2, value=count_info['misbar'])
        # closest_neighbour
        r_values = [(key,value['r_value']) for key,value in compare_dict[sample_name].items()]
        r_values.sort(key=lambda x: x[1],reverse=True)
        closest_outside_group = [value for value in r_values if value[0] not in merge_group][0]
        sheet.cell(column=8, row=row+2, value="{} with r^2={}".format(
            closest_outside_group[0],
            round(closest_outside_group[1],3)
        ))


def write_sheet_locus_info(workbook,loci_dict,loci_filters):
    sheet = workbook.create_sheet('loci_info')
    header = ['loci','main_snp','containing_snp','filters']
    for i,lable in enumerate(header):
        sheet.cell(column=i+1, row=1,value=lable)
    for i,item in enumerate(loci_dict.items()):
        # adding locus
        sheet.cell(column=1, row=i+2, value=item[0])
        # adding info
        for value in item[1]:
            if isinstance(value, list):
                # all the snp
                sheet.cell(column=3, row=i+2, value=','.join(value))
                # filters off them, same order
                sheet.cell(column=4, row=i+2, value=','.join([loci_filters[x] for x in value]))
            else:
                # the main snp
                sheet.cell(column=2, row=i+2, value=value)



def write_sheet_locus_compare(workbook,loci_counts,sample_names):
    sheet = workbook.create_sheet('mapping_x_sample')
    labels = sorted(list(loci_counts.keys()))
    for label in labels:
        if 'other' in label[-5:]:
            # if there were not enough reads on a expected locus this will be added to the other group, also in the label
            labels.insert(0, labels.pop(labels.index(label)))
    # using sample_names, to make sure it's alway's in the same row
    # header
    for col, locus in enumerate(labels):
        _ = sheet.cell(column= col+2, row=1, value=locus)
    for row, sample in enumerate(sample_names):
        _ = sheet.cell(column=1, row=row + 2, value=sample)
    # values
    for col, locus in enumerate(labels):
        for row, sample in enumerate(sample_names):
            val_dict = loci_counts[locus][sample]
            _ = sheet.cell(column=col+2, row=row + 2, value="{}".format(val_dict))


def write_sheet_genotype_compare(workbook,compare_dict,labels=None):
    sheet = workbook.create_sheet('sample_x_sample')
    if not labels:
        labels = compare_dict.keys()
    # header
    for col, sample1 in enumerate(labels):
        _ = sheet.cell(column=col + 2, row=1, value=sample1)
    # values
    for row, sample2 in enumerate(labels):
        _ = sheet.cell(column=1, row=row + 2, value=sample2)
        for col, sample1 in enumerate(labels):
            _ = sheet.cell(column=col + 2, row=row + 2, value="{}".format({
                key:round(value,3) for key,value in compare_dict[sample2][sample1].items()
            }))


def main(args):
    #TODO data is from one specific species, species obtained from vcf name
    species = args.vcf_file.split('.')[1]
    # opening the vcf file
    vcf_reader = vcf.Reader(filename=args.vcf_file)
    # opening the gff
    gff3 = gff3_manager.gff3_file(args.gff3)
    # barcode file
    rg_info = obtain_barcode_info_xlsx(args.info)
    # get the relevant sample names
    sample_id_list = []
    samples_to_merge = {}
    for sample in rg_info:
        if sample['Sample_name'] in vcf_reader.samples and sample['Species'].lower() == species.lower():
            sample_id_list.append(sample['Sample_name'])
            if sample['Cultivar_name'] in samples_to_merge:
                samples_to_merge[sample['Cultivar_name']].append(sample['Sample_name'])
            else:
                samples_to_merge[sample['Cultivar_name']] = [sample['Sample_name']]
    # looping on the loci and gathering data
    snp_dict = {sample:{} for sample in sample_id_list}
    loci_dict = {}
    snp_list = []
    loci_filters = {}
    for snp in vcf_reader:
        site_id = snp.CHROM + '_{:09}'.format(snp.POS)
        snp_list.append(site_id)
        loci_filters[site_id] = ', '.join(snp.FILTER)
        ad_fractions_dict = summeryFigures_allel_calculation(snp)
        for key,value in ad_fractions_dict.items():
            if key in sample_id_list:
                snp_dict[key][site_id] = value
        # and adding the snp to locus info
        gff_set = gff3.set_containing_position(snp.CHROM, snp.POS)
        if gff_set:
            locus_id = gff_set.pcr_product[0].attributes.id
            if locus_id in loci_dict:
                loci_dict[locus_id][0].append(site_id)
            else:
                main_snp = gff_set.get_main_snp()
                main_snp_id = main_snp.seqid + '_{:09}'.format(main_snp.start)
                loci_dict[locus_id] = [[site_id],main_snp_id]
        else:
            "for some reason this snp was not part of a locus"
    # counts
    """     this is for the counts not split by location
    counts = {}
    with open(args.counts) as count_file:
        counts_header = count_file.readline().split()
        for line in count_file:
            splitline = line.split()
            sample_name = '_'.join(splitline[0].split('_')[2:])
            counts[sample_name] = {counts_header[i]:value for i,value in enumerate(splitline)}
    """
    # expecting the counts file split by location (.xlsx)
    raw_counts = obtain_counts_xlsx(args.counts,sample_id_list)
    counts = combine_count_totals(raw_counts)
    # info by location
    loci_counts = combine_count_positions(raw_counts, gff3)
    # create pairwise comparison
    compare_dict = compare_sample_calls(snp_list,sample_id_list,snp_dict)
    # preparing the fractions file.
    wb_out = Workbook()
    # create the main sample info sheet
    write_sheet_main(wb_out,sample_id_list,rg_info,compare_dict,samples_to_merge,counts)
    # create locus information sheet
    write_sheet_locus_info(wb_out,loci_dict,loci_filters)
    # create the genotype sheet
    create_sheet_genotype_data(wb_out,snp_dict,sample_id_list,snp_list)
    # create the sample compare sheet
    write_sheet_genotype_compare(wb_out,compare_dict,sample_id_list)
    # create locus comparison sheet
    write_sheet_locus_compare(wb_out,loci_counts,sample_id_list)
    # save the excel file.
    wb_out.save(args.outfile)


if __name__ == '__main__':
    args = argument_parse()
    main(args)