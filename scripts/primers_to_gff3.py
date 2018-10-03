"""
Throw away script for creating a gff3 with already designed primers
"""
import sys
from argparse import ArgumentParser
from openpyxl import load_workbook
from pyfasta import Fasta

illumina_FWD = "CGACAGGTTCAGAGTTCTACAGTCCGACGATC"
illumina_REV = "GTGACTGGAGTTCAGACGTGTGCTCTTCCGATCT"


def argument_parser():
    parser = ArgumentParser(description='Options for fastq RG tagger')
    parser.add_argument(
        '-g', '--gff3',
        help="the gff3 file writen with the primers as gff3 features"
    )
    parser.add_argument(
        '-i', '--info',
        required=True,
        help="the .xlsx file with primers for each SNP"
    )
    parser.add_argument(
        '-f', '--fasta',
        dest="ref_seq",
        required=True,
        help="the ref genome the primers were designed on."
    )
    parser.add_argument(
        '-b', '--base',
        required=True,
        help="the gff3 file the primers will be added to"
    )
    return parser.parse_args()


def extract_gff3_fields(primer_info):
    """
    The gff3 file is a format to annotate locations on a genome.
    The gff3 annotations requires locations on the genome, this was not required for creating the primers.
    This function computes these locations and prepares the data for the gff3 file
    :param primer_info: the sequence and target used for creating primers
    :return: lines to be writen in gff3 file
    """
    gff3_objects = []
    gt_seq_region = primer_info['ref']+'_'+str(primer_info['forward_start'])+'-'+str(primer_info['reverse_stop'])
    SNP_name = primer_info['ref'] + '_' + str(primer_info['pos'])
    # creating the parent
    gff3_objects.append({
        'seqid' : primer_info['ref'],
        'source' : "Primer3",
        'type' : "region",
        'start': str(primer_info['forward_start']),
        'end' : str(primer_info['reverse_stop']),
        'score' : '.',
        'strand' : '+',
        'phase' : '.', # used for coding DNA sequence (CDS)
        'attributes' : {
            'ID' : gt_seq_region,
            'Name' : primer_info['ref']+'_'+str(primer_info['pos']),
            'Parent' : primer_info['parent'],
            'Note' : "GT-seq target region"
        }
    })
    gff3_objects.append({
        'seqid' : primer_info['ref'],
        'source' : "Primer3",
        'type' : "PCR_product",
        'start': str(primer_info['forward_start']),
        'end' : str(primer_info['reverse_stop']),
        'score' : '.',
        'strand' : '+',
        'phase' : '.', # used for coding DNA sequence (CDS)
        'attributes' : {
            'ID' : gt_seq_region + "_pcr",
            'Name' : SNP_name + " PCR product",
            'Parent' : gt_seq_region,
            'Note' : "GT-seq primer product"
        }
    })
    #forward primer
    gff3_objects.append({
        'seqid': primer_info['ref'],
        'source': "Primer3",
        'type': "forward_primer",
        'start': str(primer_info['forward_start']),
        'end': str(primer_info['forward_stop']),
        'score': '.',
        'strand': '+',
        'phase': '.',  # used for coding DNA sequence (CDS)
        'attributes': {
            'ID': gt_seq_region + "_FW",
            'Name': SNP_name+"_FW",
            'Parent': gt_seq_region
        }
    })
    #reverse primer
    gff3_objects.append({
        'seqid': primer_info['ref'],
        'source': "Primer3",
        'type': "reverse_primer",
        'start': str(primer_info['reverse_start']),
        'end': str(primer_info['reverse_stop']),
        'score': '.',
        'strand': '-',
        'phase': '.',  # used for coding DNA sequence (CDS)
        'attributes': {
            'ID': gt_seq_region + "_RV",
            'Name': SNP_name+"_RV",
            'Parent': gt_seq_region
        }
    })
    return gff3_objects


def obtain_barcode_info_xlsx(info_file):
    info_lines = []
    wb = load_workbook(info_file, data_only=True)
    if 'Library' in wb.sheetnames:
        info_sheet = wb['Library']
    else:
        info_sheet = wb.active
    info_headers = []
    for column in info_sheet.iter_cols(max_row=1):
        info_headers.append(column[0].value)
    for row in info_sheet.iter_rows(min_row=2):
        if row[3]:
            info_dict = {info_headers[i]:item.value for i,item in enumerate(row) if item.value}
            info_lines.append(info_dict)
    return info_lines


def reverse_complement(dna):
    complement = {'A': 'T', 'C': 'G', 'G': 'C', 'T': 'A'}
    return ''.join([complement[base] for base in dna[::-1]])


def write_gff3(gff3_file, primer_data,f=None):
    gff3_file.write("##gff-version 3\n")
    sorted_keys = sorted(primer_data.keys())
    last_region = ""
    for key in sorted_keys:
        if f and key.split('_')[0] != last_region:
            last_region = key.split('_')[0]
            seq_region = f[last_region]
            gff3_file.write("##sequence-region {} {} {}\n".format(last_region, 1, (seq_region.stop + 1)))
        gff3_lines = extract_gff3_fields(primer_data[key])
        for line in gff3_lines:
            line['attributes'] = ';'.join([key + '=' + value for key, value in line['attributes'].items()])
            keys = ['seqid', 'source', 'type', 'start', 'end', 'score', 'strand', 'phase', 'attributes']
            gff3_file.write('\t'.join([line[key] for key in keys]) + '\n')


def main(args):
    if args.gff3:
        outfile = open(args.gff3, 'w')
    else:
        outfile = sys.stdout

    primers = {}

    f = Fasta(args.ref_seq)

    with open(args.info, 'br') as info_file:
        info_lines = obtain_barcode_info_xlsx(info_file)
        for line in info_lines:
            flank_start = line['position']-99
            flank_stop = line['position']+101
            flank_seq = f[line['Name']][flank_start:flank_stop]

            primer_info = {}
            id = line['Name']+'_'+str(flank_start)+'-'+str(flank_stop)
            primer_info['parent'] = id
            primer_info['flank_start'] = flank_start
            primer_info['flank_stop'] = flank_stop
            primer_info['flank_seq'] = flank_seq
            primer_info['ref'] = line['Name']
            primer_info['pos'] = line['position']
            primer_info['forward_unique'] = line['FWD_Primer'][len(illumina_FWD):]
            primer_info['reverse_unique'] = reverse_complement(line['REV_Primer'][len(illumina_REV):])
            primer_info['forward_start'] = flank_seq.index(primer_info['forward_unique'])+flank_start
            primer_info['forward_stop'] = primer_info['forward_start']+len(primer_info['forward_unique'])
            primer_info['reverse_start'] = flank_seq.index(primer_info['reverse_unique'])+flank_start
            primer_info['reverse_stop'] = primer_info['reverse_start']+len(primer_info['reverse_unique'])
            primers[id] = primer_info
    write_gff3(outfile,primers,f)
    outfile.close()


if __name__ == '__main__':
    args = argument_parser()
    main(args)