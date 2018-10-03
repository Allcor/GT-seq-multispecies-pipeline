#!/usr/local/bin/python3.6
"""
read a vcf file and get the allel persentage for all loci and sample combinations
"""

from openpyxl import load_workbook
from collections import defaultdict
from argparse import ArgumentParser, FileType
from plotly import tools
import numpy
import scipy
from scipy import optimize, stats
import pandas as pd
import colorlover as cl
import plotly.offline as pyoff
import plotly.plotly as py
import plotly.graph_objs as go
import plotly.figure_factory as ff
import sys
import vcf

html_template = """
<html>
<!-- Arlo Hoogeveen, Naktuinbouw 2018 -->
<head>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.1/css/bootstrap.min.css">
    <style>body{{ margin:0 100; background:whitesmoke; }}</style>
</head>
<body>
  <!-- Output from Plotly Python script -->
  <h1>gt_seq replicate samples data.</h1>
  {plot}
</body>
</html>
"""

py.sign_in('ThomasvanGurp','X1kwzANshwWjiWdlmKBR')
colors = ['rgba(230, 25, 75, .8)',
          'rgba(60, 180, 75, .8)',
          'rgba(255, 225, 25, .8)',
          'rgba(0, 130, 200, .8)',
          'rgba(245, 130, 48, .8)',
          'rgba(145, 30, 180, .8)',
          'rgba(70, 240, 240, .8)',
          'rgba(240, 50, 230, .8)',
          'rgba(210, 245, 60, .8)',
          'rgba(250, 190, 190, .8)',
          'rgba(0, 128, 128, .8)',
          'rgba(230, 190, 255, .8)',
          'rgba(170, 110, 40, .8)',
          'rgba(255, 250, 200, .8)',
          'rgba(128, 0, 0, .8)',
          'rgba(170, 255, 195, .8)',
          'rgba(128, 128, 0, .8)',
          'rgba(255, 215, 180, .8)',
          'rgba(0, 0, 128, .8)',
          'rgba(128, 128, 128, .8)',
          ]

colors2 = [ 'rgba(213, 255, 0, .7)',
            'rgba(255, 0, 86, .7)',
            'rgba(158, 0, 142, .7)',
            'rgba(14, 76, 161, .7)',
            'rgba(255, 229, 2, .7)',
            'rgba(0, 95, 57, .7)',
            'rgba(0, 255, 0, .7)',
            'rgba(149, 0, 58, .7)',
            'rgba(255, 147, 126, .7)',
            'rgba(164, 36, 0, .7)',
            'rgba(0, 21, 68, .7)',
            'rgba(145, 208, 203, .7)',
            'rgba(98, 14, 0, .7)',
            'rgba(107, 104, 130, .7)',
            'rgba(0, 0, 255, .7)',
            'rgba(0, 125, 181, .7)',
            'rgba(106, 130, 108, .7)',
            'rgba(0, 174, 126, .7)',
            'rgba(194, 140, 159, .7)',
            'rgba(190, 153, 112, .7)',
            'rgba(0, 143, 156, .7)',
            'rgba(95, 173, 78, .7)',
            'rgba(255, 0, 0, .7)',
            'rgba(255, 0, 246, .7)',
            'rgba(255, 2, 157, .7)',
            'rgba(104, 61, 59, .7)',
            'rgba(255, 116, 163, .7)',
            'rgba(150, 138, 232, .7)',
            'rgba(152, 255, 82, .7)',
            'rgba(167, 87, 64, .7)',
            'rgba(1, 255, 254, .7)',
            'rgba(255, 238, 232, .7)',
            'rgba(254, 137, 0, .7)',
            'rgba(189, 198, 255, .7)',
            'rgba(1, 208, 255, .7)',
            'rgba(187, 136, 0, .7)',
            'rgba(117, 68, 177, .7)',
            'rgba(165, 255, 210, .7)',
            'rgba(255, 166, 254, .7)',
            'rgba(119, 77, 0, .7)',
            'rgba(122, 71, 130, .7)',
            'rgba(38, 52, 0, .7)',
            'rgba(0, 71, 84, .7)',
            'rgba(67, 0, 44, .7)',
            'rgba(181, 0, 255, .7)',
            'rgba(255, 177, 103, .7)',
            'rgba(255, 219, 102, .7)',
            'rgba(144, 251, 146, .7)',
            'rgba(126, 45, 210, .7)',
            'rgba(189, 211, 147, .7)',
            'rgba(229, 111, 254, .7)',
            'rgba(222, 255, 116, .7)',
            'rgba(0, 255, 120, .7)',
            'rgba(0, 155, 255, .7)',
            'rgba(0, 100, 1, .7)',
            'rgba(0, 118, 255, .7)',
            'rgba(133, 169, 0, .7)',
            'rgba(0, 185, 23, .7)',
            'rgba(120, 130, 49, .7)',
            'rgba(0, 255, 198, .7)',
            'rgba(255, 110, 65, .7)',
            'rgba(232, 94, 190, .7)',
            'rgba(1, 0, 103, .7)',
            'rgba(0, 0, 0, .7)']

def argument_parse():
    parser = ArgumentParser(description='Options for BarcodeSplit script')
    parser.add_argument(
        "-v", "--vcf",
        dest="vcf_file",
        required=True,
        help="the path to the vcf file containing varients of the GT-seq run."
    )
    parser.add_argument(
        '-i', '--info',
        help="the .csv or .xlsx file with information on the read groups"
    )
    parser.add_argument(
        '-o', '--outfile',
        nargs='?',
        type=FileType('w'),
        default = sys.stdout
    )
    parser.add_argument(
        '-c', '--counts',
        help="the file containing the read counts"
    )
    parser.add_argument(
        '-p', '--plots',
        help="The path of the folder to write the plots to."
    )
    parser.add_argument(
        '-g', '--grid',
        action = 'store_false',
        help="Suppresses the message for the plotly create_subplots grid"
    )
    return parser.parse_args()


def summeryFigures_allel_calculation(loci):
    sample_dict = {}
    for sample in loci.samples:
        values = {}
        # size = len(info)
        # if size == 13:
        #    A1_corr = float(info[6])
        #    A2_corr = float(info[7])
        # xarr = info[1].split('=')
        if not sample.gt_nums or sample.gt_alleles[0] == sample.gt_alleles[1]:
            allel_x = 0
            if sample.gt_alleles[0] and sample.gt_alleles[0] != '0':
                allel_y = int(sample.gt_alleles[0])
            else:
                allel_y = 1
        else:
            allel_x, allel_y = [int(i) for i in sample.gt_nums.split('/')]
        if isinstance(sample.data.AD, list):
            xarr = [loci.alleles[allel_x], sample.data.AD[allel_x]]
            try:
                yarr = [loci.alleles[allel_y].sequence, sample.data.AD[allel_y]]
            except AttributeError:
                yarr = [loci.alleles[allel_y].type, sample.data.AD[allel_y]]
        else:
            xarr = [loci.alleles[allel_x], sample.data.AD]
            yarr = ['.', 0]

        # yarr = info[2].split('=')
        values['x'] = int(round(float(xarr[1])))
        values['y'] = int(round(float(yarr[1])))
        values['genotype'] = sample.gt_nums
        AF_div = float(xarr[1]) + float(yarr[1])
        if AF_div == 0:
            AF_div = 0.1
        values['p_A2'] = (float(yarr[1]) / AF_div) * 100
        sample_dict[sample.sample] = values
    return sample_dict


def obtain_barcode_info_xlsx(info_file):
    info_lines = []
    wb = load_workbook(info_file, data_only=True)
    if 'Library' in wb.sheetnames:
        info_sheet = wb['Library']
    else:
        info_sheet = wb.active
    info_headers = []
    for column in info_sheet.iter_cols(max_row=1):
        info_headers.append(column[0].value)
    for row in info_sheet.iter_rows(min_row=2):
        info_dict = {info_headers[i]:item.value for i,item in enumerate(row)}
        # rows with empty cells are not allowed
        if all(x for x in info_dict.values()):
            info_lines.append(info_dict)
    return info_lines


def make_plot(name,set,site_list,loci_dict,filename,counts=None):
    allel_data = []
    layout = go.Layout(
        title='sample calls {}'.format(name),
        width=1500,
        height=640,
        legend=dict(
            orientation='h'
        ),
        yaxis=dict(
            range=[0, 100],
            title = 'percentige of allele'
        ),
        xaxis=dict(
            showticklabels=False,
            title = 'Loci'
        )
    )
    for i,sample in enumerate(set):
        if counts and sample in counts:
            sample_label = str(sample) + " - RD = {}".format(counts[sample])
        else:
            sample_label = sample
        color = colors2[i]
        sample_trace = go.Scatter(
            x = site_list,
            y = [loci_dict[str(sample)][x]['p_A2'] for x in site_list],
            name = sample_label,
            mode='markers',
            marker=dict(
                size=5,
                color=color,
                line=dict(
                    width=0.5,
                )
            )
        )
        allel_data.append(sample_trace)

    #write the file
    fig = go.Figure(data=allel_data, layout=layout)
    py.image.save_as(fig, filename=filename)

def make_trace_genotype(sample,loci_dict,site_list,color,sample_label=None):
        if not sample_label:
            sample_label = sample
        y_data = []
        for locus in site_list:
            site = loci_dict[str(sample)][locus]
            if site['genotype'] == '0/0':
                y_data.append(site['p_A2'])
            if site['genotype'] == '0/1':
                y_data.append(site['p_A2']-50)
            if site['genotype'] == '1/1':
                y_data.append(site['p_A2']-100)
        #if counts and sample in counts:
        #    sample_label = str(sample) + " RD = {}".format(counts[sample])
        #else:
        #    sample_label = sample
        sample_trace = go.Box(
            name = sample_label,
            y = y_data,
            boxpoints='all',
            marker = dict(
                color = color),
            line = dict(
                color = color)
        )
        return sample_trace

def make_trace_average(sample,loci_dict,site_list,color,averages,sample_label=None):
    if not sample_label:
        sample_label = sample
    y_data = []
    for locus in site_list:
        site = loci_dict[str(sample)][locus]
        y_data.append(site['p_A2'] - averages[locus])
    # if counts and sample in counts:
    #    sample_label = str(sample) + " RD = {}".format(counts[sample])
    # else:
    #    sample_label = sample
    sample_trace = go.Box(
        name=sample_label,
        y=y_data,
        boxpoints='all',
        showlegend = False,
        marker=dict(
            color=color),
        line=dict(
            color=color)
    )
    return sample_trace


def make_compare_trace(sample_x,sample_y,loci_list,loci_dict):
    x = []
    y = []
    marker_color = []
    marker_symbol = []
    marker_text = []
    different_count = 0
    for locus in loci_list:
        site_x = loci_dict[str(sample_x)][locus]
        x.append(round(site_x['p_A2'],1))
        site_y = loci_dict[str(sample_y)][locus]
        y.append(round(site_y['p_A2'],1))
        if site_x['genotype']:
            # if a other base is called then the first alt, other colors must be used.
            #color_nr = sum([int(i) for i in site_x['genotype'] if i.isdigit()])
            # this would be the same color for 0/2 as for 1/1
            geno = [int(i) for i in site_x['genotype'] if i.isdigit()]
            max_geno = max(geno)-1
            offset = int((max_geno * (max_geno+1))/2) # using a Triangular Number Sequence as offset
            # each time a different allel is added the offset is dependent by the
            # amount of combinations that can be made with alleles before
            color_nr = sum(geno)+offset
            marker_color.append(colors[color_nr])
        else:
            marker_color.append('gray')
        if site_y['genotype'] == site_x['genotype']:
            marker_symbol.append('circle')
        else:
            marker_symbol.append('x')
            different_count += 1
        marker_text.append(locus+"<br>AD_x: [{x1},{y1}]<br>AD_y: [{x2},{y2}]".format(
            x1=site_x['x'],
            y1=site_x['y'],
            x2=site_y['x'],
            y2=site_y['y'],
        ))
    #create the linear fit
    slope, intercept, r_value, p_value, std_err = scipy.stats.linregress(x, y)
    layout_text = "R<sup>^2</sup> = {r_value:.2f},<br>Y = {slope:.2f}X + {intercept:.2f}<br>{off}/{totalloci} ({off_percent}%) calls off".format(
        r_value=r_value,
        slope=slope,
        intercept=intercept,
        off=different_count,
        totalloci=len(loci_list),
        off_percent=int(((different_count/len(loci_list))*100)+0.5),
    )
    xi = numpy.arange(0, 110, 10)
    line = slope * xi + intercept
    fit_trace = go.Scatter(
        x=xi,
        y=line,
        mode='lines',
        marker=go.Marker(color='rgb(31, 119, 180)'),
        name='Fit'
    )
    # addding the r^2 to the sample info
    if 'r_squared' in loci_dict[str(sample_x)]['info']:
        loci_dict[str(sample_x)]['info']['r_squared'].append(r_value)
    else:
        loci_dict[str(sample_x)]['info']['r_squared'] = [r_value]
    #create the plot
    compare_trace = go.Scatter(
        name=sample_x,
        x = x,
        y = y,
        mode='markers',
        text=marker_text,
        marker=dict(
            color=marker_color,
            symbol=marker_symbol,
            size=5,
            line=dict(
                width=0.5,
            )
        )
    )
    return compare_trace, layout_text, fit_trace


def make_dendogram_plot(name,rep_set,loci_names,loci_dict,filename):
    dendogram_matrix = []
    for sample1 in rep_set:
        dendogram_array = []
        for sample2 in rep_set:
            correct_calls = []
            for locus in loci_names:
                if loci_dict[sample1][locus]['genotype'] == loci_dict[sample2][locus]['genotype']:
                    correct_calls.append(1)
                else:
                    correct_calls.append(0)
            cell_percent = sum(correct_calls) / len(correct_calls)
            dendogram_array.append(cell_percent)
        dendogram_matrix.append(dendogram_array)
    data_frame = numpy.array(dendogram_matrix)
    dendro = ff.create_dendrogram(data_frame, orientation='left', labels=rep_set)
    dendro['layout'].update({'width':800, 'height':500})
    py.image.save_as(dendro, filename=filename)


def make_table_plot(name,set,site_list,loci_dict,filename):
    ryg = cl.scales['11']['div']['RdYlGn']
    ryg100 = cl.interp(ryg, 100)
    table_values = [set]
    colors = [['grey' for i in set]]
    for sample1 in set:
        table_line = []
        table_color = []
        for sample2 in set:
            correct_calls = []
            if sample1 == sample2:
                cell_percent = 'X'
                table_line.append(cell_percent)
                cell_color = 'grey'
            else:
                for locus in site_list:
                    if loci_dict[sample1][locus]['genotype'] == loci_dict[sample2][locus]['genotype']:
                        correct_calls.append(1)
                    else:
                        correct_calls.append(0)
                cell_percent = int((sum(correct_calls)/len(correct_calls))*100)
                cell_color = ryg100[cell_percent-1]
                table_line.append(str(cell_percent)+'%')
            table_color.append(cell_color)
        table_values.append(table_line)
        colors.append(table_color)
    trace = go.Table(
        header = dict(
            values=['']+set,
            line = dict(color = '#506784'),
            fill=dict(color='grey'),
            font = dict(color = 'white', size = 12)
        ),
        cells = dict(
            values=table_values,
            line = dict(color = '#506784'),
            fill = dict(color = colors),
            font=dict(color=['white','lightgrey'], size=12)
        )
    )
    layout = go.Layout(
        title='percentage of same calls {}'.format(name),
        width=1500,
        height=640
    )
    fig = dict(data=[trace], layout=layout)
    py.image.save_as(fig, filename=filename)


def make_html_table(name,set,site_list,loci_dict):
    # collect the fraction of same calls
    table_values = []
    for sample1 in set:
        table_line = []
        for sample2 in set:
            correct_calls = []
            for locus in site_list:
                if loci_dict[sample1][locus]['genotype'] == loci_dict[sample2][locus]['genotype']:
                    correct_calls.append(1)
                else:
                    correct_calls.append(0)
            cell_percent = sum(correct_calls) / len(correct_calls)
            table_line.append(cell_percent)
        table_values.append(table_line)
    # and now make the data frame
    compare_frame = pd.DataFrame( table_values ).transpose()
    compare_frame.rename(
        columns={i:name for i,name in enumerate(set)},
        index={i:name for i,name in enumerate(set)},
        inplace=True
    )
    """
    compare_table = compare_frame.round(3).style.background_gradient(
        cmap='viridis_r',
        high=1.5,
        low=0
    ).highlight_max(axis=0,color='lightgray').render()
    #print(compare_table)
    compare_table = compare_table.replace(
        '<table ',
        '<table class="table table-striped" '
    )
    """
    compare_table = compare_frame.to_html().replace(
        '<table border="1" class="dataframe">',
        '<table class="table table-striped">'
    )
    return compare_table


def make_box_plot(name,set,site_list,loci_dict,filename,counts=None):
    fig = tools.make_subplots(
        rows=2,
        cols=1,
        subplot_titles=('heterozygous around 50%', 'heterozygous around average')
    )
    fig['layout'].update(
        title='sample calls {}'.format(name),
        width=640,
        height=800,
        legend=dict(
            orientation='h'
        )
    )
    fig['layout']['xaxis1'].update(
        showticklabels=False,
        title='samples'
    )
    fig['layout']['xaxis2'].update(
        showticklabels=False,
        title='samples'
    )
    fig['layout']['yaxis1'].update(
        range=[-50, 50],
        title = 'error of hetrozygotes'
    )
    fig['layout']['yaxis2'].update(
        range=[-50, 50],
        title = 'error of hetrozygotes'
    )
    # normalise data on called allele.
    for i,sample in enumerate(set):
        color = colors2[i]
        sample_trace = make_trace_genotype(sample,loci_dict,site_list,color)
        fig.append_trace(sample_trace,1,1)
    # normalise on average
    # create average of alleles for this set
    averages = {}
    for locus in site_list:
        values = []
        for sample in set:
            values.append(loci_dict[str(sample)][locus]['p_A2'])
        averages[locus] = sum(values)/len(values)
    # create all the boxplots
    for i,sample in enumerate(set):
        color = colors2[i]
        sample_trace = make_trace_average(sample,loci_dict,site_list,color,averages)
        fig.append_trace(sample_trace,2,1)
    #write the file
    py.image.save_as(fig, filename=filename)


def make_comparison_plots(name,set,loci_list,loci_dict,counts=None,subplots=True):
    set_size = len(set)
    annotations = []
    height = set_size*300
    width = set_size*300
    fig = tools.make_subplots(
        print_grid=subplots,
        rows=set_size,
        cols=set_size,
    )
    # horizontal_spacing (kwarg, float in [0,1], default=0.2 / columns)
    horizontal_spacing = 0.2/set_size
    # vertical_spacing (kwarg, float in [0,1], default=0.3 / rows)
    vertical_spacing = 0.3/set_size
    for plot_column in range(1,set_size+1):
        colory = colors2[plot_column]
        for plot_row in range(1,set_size+1):
            colorx = colors2[plot_row]
            plot_nr = plot_column + (plot_row - 1) * set_size
            if plot_row == plot_column:
                """
                if counts:
                    plot_nr = plot_x+(plot_y-1)*set_size
                    test_domain = dict(
                        x = fig['layout']['xaxis{}'.format(plot_nr)]['domain'],
                        y = fig['layout']['yaxis{}'.format(plot_nr)]['domain']
                    )
                    test_dict = go.Table(
                        columnwidth = [10,30],
                        domain = test_domain,
                        header = dict(
                            values = ['', set[plot_x-1]],
                            font = dict(size = 8),
                        ),
                        cells = dict(
                            values = ['rc', counts[set[plot_x-1]]],
                        )
                    )
                    fig['data'].append(test_dict)
                else:
                    pass
                """
                if counts:
                    info_text = "x-axis of this row: <br>{} <br> <br>pos: {},{} <br>read count: {} ".format(
                        set[plot_column-1],
                        loci_dict[set[plot_column-1]]['info']['plate'],
                        loci_dict[set[plot_column-1]]['info']['position'],
                        counts[set[plot_column-1]]
                    )
                    domain_x = fig['layout']['xaxis{}'.format(plot_nr)]['domain']
                    domain_y = fig['layout']['yaxis{}'.format(plot_nr)]['domain']
                    offset = (0.05 / (set_size-1))
                    annotation = go.Annotation(
                        x=domain_x[0],
                        y=domain_y[1],
                        width=((domain_x[1]-domain_x[0])-offset)*width,
                        height=(domain_y[1]-domain_y[0])*height,
                        xref='paper',
                        yref='paper',
                        yanchor='top',
                        xanchor='left',
                        text=info_text,
                        align='right',
                        showarrow=False,
                        bgcolor = 'lightgray'#colorx
                    )
                    annotations.append(annotation)
            #elif plot_x > plot_y:
                # half of the grid to safe the server some work.
            #    pass
            else:
                trace, layout_text, fit_trace = make_compare_trace(set[plot_row-1],set[plot_column-1],loci_list,loci_dict)
                fig.append_trace(fit_trace,plot_row,plot_column)
                fig.append_trace(trace,plot_row,plot_column)
                fig['layout']['xaxis{}'.format(plot_nr)].update(
                    tickvals = [0, 25, 50, 75, 100],
                    ticktext = ['0%', '', '50%', '', '100%']
                )
                #tickfont= dict(color=colorx)

                fig['layout']['yaxis{}'.format(plot_nr)].update(
                    tickvals = [0, 25, 50, 75, 100],
                    ticktext = ['0%','','50%','','100%'],
                )
                #tickfont = dict(color=colory)
                offset = (0.05/set_size)
                # x = 20,
                # y = 90,
                # xref = 'x' + str(plot_nr),
                # yref = 'y' + str(plot_nr),
                annotation = go.Annotation(
                    x = fig['layout']['xaxis{}'.format(plot_nr)]['domain'][0]+offset,
                    y = fig['layout']['yaxis{}'.format(plot_nr)]['domain'][1],
                    xref = 'paper',
                    yref = 'paper',
                    yanchor = 'middle',
                    xanchor = 'left',
                    text=layout_text,
                    align='left',
                    font=go.Font(size=8),
                    showarrow=False
                )
                annotations.append(annotation)
    # fix the layout
    # default figure margins: L=80,R=80,T=100,B=80
    fig['layout'].update(
        title='proportion comparison {}'.format(name),
        width=width+160,
        height=height+180,
        showlegend=False,
        hovermode='closest',
        legend=dict(
            orientation='h'
        ),
        annotations = annotations
    )
    # write the file
    #py.image.save_as(fig, filename=filename)
    div = pyoff.plot(fig, include_plotlyjs=False, output_type='div')
    return div


def make_readdepth_regression_plot(loci_dict,average=True):
    # this requires the R^2 added to loci_dict in the 'make_comparison_plots' function
    layout = go.Layout(
        title='distribution of the regression fits',
        width=1500,
        height=640,
        hovermode='closest',
        legend=dict(
            orientation='h'
        ),
        yaxis=dict(
            title='R^2 value'
        ),
        xaxis=dict(
            title='total reads of '
        )
    )
    r_list = []
    c_list = []
    t_list = []
    for sample,value in loci_dict.items():
        info = value['info']
        if 'r_squared' in info:
            if average:
                r_list.append(sum(info['r_squared'])/len(info['r_squared']))
            else:
                r_list.append(max(info['r_squared']))
            c_list.append(int(info['count']))
            t_list.append(sample)
    #scatter trace
    scatter = go.Scatter(
        x = c_list,
        y = r_list,
        mode='markers',
        text=t_list,
        marker=dict(
            size=5,
            line=dict(
                width=0.5,
            )
        )
    )
    # fitting logaritmic function
    """
    #creating subset
    c_sub_list = []
    r_sub_list = []
    c_treshold = 300000
    for i,count in enumerate(c_list):
        if count < c_treshold:
            c_sub_list.append(count)
            r_sub_list.append(r_list[i].round(2))
    """
    # the math function as python function
    def fivepl(x, b, c):
        a = 0
        d = 1
        g = 0.25
        # https: // en.wikipedia.org / wiki / Generalised_logistic_function
        # https://stats.stackexchange.com/questions/190107/curve-fit-with-logarithmic-regression-in-python/338472#338472
        return (((a - d) / numpy.power(1 + (numpy.power((x / c),b)), g)) + d)
    #popt,pcov = scipy.optimize.curve_fit(lambda t,a,b: a+b*numpy.log(t), c_list, r_list)
    popt,pcov = scipy.optimize.curve_fit(fivepl, c_list, r_list, bounds=([0.1,1000],[5,10000]))
    fit_max = max(c_list)
    xi = numpy.arange(1, fit_max, (fit_max/100))
    line = fivepl(xi,*popt)
    fit_trace = go.Scatter(
        x=xi,
        y=line,
        mode='lines',
        marker=go.Marker(color='rgb(31, 119, 180)'),
        name='Fit'
    )
    layout_text = "growth rate = {:.2f}<br>inflection point = {:0f}".format(*popt)
    annotation1 = go.Annotation(
        x=0,
        y=1,
        xref='paper',
        yref='paper',
        yanchor='middle',
        xanchor='left',
        text=layout_text,
        align='left',
        font=go.Font(size=8),
        showarrow=False
    )
    #get the count cutoff
    cutoff = optimize.fsolve(lambda x: fivepl(x, *popt) - 0.965,1000)
    annotation2 = go.Annotation(
        x=cutoff[0],
        y=0.5,
        yanchor="middle",
        xanchor='left',
        text=str(int(cutoff[0])),
        align='left',
        showarrow=False
    )
    layout.update(
        shapes=[{
            'type': 'line',
            'x0': cutoff[0],
            'x1': cutoff[0]+1,
            'y0':0,
            'y1':1,
            'line': {
                'color': 'rgb(55, 128, 191)',
                'width': 3,
                'dash': 'dashdot',
            },
        }],
        annotations = [annotation1,annotation2]
    )
    # return the graph
    fig = dict(data=[scatter,fit_trace], layout=layout)
    div = pyoff.plot(fig, include_plotlyjs=False, output_type='div')
    return div


def make_all_plots(name,set,site_list,loci_dict,filename,counts=None):
    # get only the hetrozygotes
    # site_list_hz = [locus for locus in site_list if loci_dict[set[0]][locus]['genotype'] == '0/1']
    # initiate plot
    fig = tools.make_subplots(
        rows=2,
        cols=2,
        specs=[[{'rowspan': 2},{}],
               [None,{}]],
        subplot_titles=('scater plot', 'error from call', 'error from average')
    )
    fig['layout'].update(
        title='sample calls {}'.format(name),
        width=1200,
        height=800,
        hovermode='closest',
        legend=dict(
            orientation='h'
        )
    )
    fig['layout']['xaxis2'].update(
        showticklabels=False,
        title='samples'
    )
    fig['layout']['xaxis3'].update(
        showticklabels=False,
        title='samples'
    )
    fig['layout']['yaxis2'].update(
        range=[-50, 50],
        title = 'error of hetrozygotes'
    )
    fig['layout']['yaxis3'].update(
        range=[-50, 50],
        title = 'error of hetrozygotes'
    )
    # create average of alleles for this set
    averages = {}
    for locus in site_list:
        values = []
        for sample in set:
            values.append(loci_dict[str(sample)][locus]['p_A2'])
        averages[locus] = sum(values) / len(values)
    for i,sample in enumerate(set):
        if counts and sample in counts:
            sample_label = str(sample) + " RD = {}".format(counts[sample])
        else:
            sample_label = sample
        color = colors2[i]
        sample_trace = go.Scatter(
            x = [loci_dict[str(sample)][x]['x'] for x in site_list],
            y = [loci_dict[str(sample)][x]['y'] for x in site_list],
            name = sample_label,
            mode='markers',
            text=site_list,
            showlegend = False,
            marker=dict(
                size=5,
                color=color
            )
        )
        fig.append_trace(sample_trace,1,1)
        # boxplot with normalised data on expected for called allele.
        sample_trace = make_trace_genotype(sample, loci_dict, site_list, color, sample_label)
        fig.append_trace(sample_trace, 1, 2)
        # create boxplots around average
        sample_trace = make_trace_average(sample, loci_dict, site_list, color, averages)
        fig.append_trace(sample_trace, 2, 2)
    #write the file
    #py.image.save_as(fig, filename=filename)
    div = pyoff.plot(fig, include_plotlyjs=False, output_type='div')
    with open(filename, 'w') as outfile:
        outfile.write(html_template.format(plot=div))


def main(args):
    # opening the vcf file
    vcf_reader = vcf.Reader(filename=args.vcf_file)
    # preparing the fractions file.
    sample_id_list = vcf_reader.samples
    args.outfile.write('loci_id;'+';'.join(sample_id_list)+'\n')
    # looping on the loci and gathering data
    loci_dict = {sample:{} for sample in sample_id_list}
    loci_list = []
    for loci in vcf_reader:
        site_id = loci.CHROM + '_{:07}'.format(loci.POS)
        loci_list.append(site_id)
        ad_fractions_dict = summeryFigures_allel_calculation(loci)
        for key,value in ad_fractions_dict.items():
            loci_dict[key][site_id] = value
        args.outfile.write(site_id+';'+';'.join(str(ad_fractions_dict[x]) for x in sample_id_list)+'\n')
    # make figures
    if args.plots and args.info:
        loci_list.sort()
        # get the samples to merge.
        samples_to_merge = defaultdict(list)
        rg_info = obtain_barcode_info_xlsx(args.info)
        # get the counts
        if args.counts:
            counts = {}
            if args.counts.split('.')[-1] == 'txt':
                with open(args.counts) as count_file:
                    for line in count_file:
                        splitline = line.split()
                        sample_name = '_'.join(splitline[0].split('_')[2:])
                        counts[sample_name] = splitline[1]
            elif args.counts.split('.')[-1] == 'xlsx':
                wb = load_workbook(args.counts, data_only=True)
                sheet = wb['total_reads']
                for row in sheet.iter_rows(min_row=2):
                    sample_name = '_'.join(row[0].value.split('_')[2:])
                    counts[sample_name] = sum([int(x.value) for x in row[1:]])
            else:
                raise OSError(1, 'countfile has a unrecognised extention:', args.counts)
        else:
            counts = None
        # some additional info
        for sample in rg_info:
            #TODO: for now i take the folder name as indication what species to make plots for
            if sample['Species'].lower() == args.plots.split('/')[-2].split('_')[0].lower():
                samples_to_merge[sample['Cultivar_name']].append(sample['Sample_name'])
            # adding some info of the sample to loci_dict
            if sample['Sample_name'] in loci_dict:
                loci_dict[sample['Sample_name']]['info'] = {
                    'plate':sample['Plate_name'],
                    'position':sample['Position']
                }
                if args.counts and sample['Sample_name'] in counts:
                    loci_dict[sample['Sample_name']]['info']['count'] = counts[sample['Sample_name']]
        # make the plots
        for name,set in samples_to_merge.items():
            filename = args.plots + 'sample_{}.png'
            filename2 = args.plots + 'sample_{}.html'
            keepcharacters = ('.','_')
            name = name.replace(' ', '_')
            fix_name = "".join(c for c in name if c.isalnum() or c in keepcharacters).rstrip()
            # scatter plot and error plots
            #make_all_plots(name,set,loci_list,loci_dict,filename2.format(fix_name),counts)
            # just the error plot
            # make_box_plot(name,set,loci_list,loci_dict,filename.format(fix_name1),counts)
            # just the scatter plot
            # fix_name2 = fix_name+'_test2'
            # make_scatter_plot(name,set,loci_list,loci_dict,filename.format(fix_name2),counts)
            # just the loci plot
            #fix_name1 = fix_name+'_test'
            #make_plot(name,set,loci_list,loci_dict,filename.format(fix_name1),counts)
            # table with the percentages of each combo
            #fix_name_table = fix_name+'_table'
            #make_table_plot(name,set,loci_list,loci_dict,filename.format(fix_name_table))
            # comparison persentage plot.
            fix_name_compare = fix_name+'_compare'
            if len(set) > 1:
                compare_div = make_comparison_plots(name,set,loci_list,loci_dict,counts,args.grid)
                html_table = make_html_table(name,set,loci_list,loci_dict)
                with open(filename2.format(fix_name_compare), 'w') as outfile:
                    outfile.write(html_template.format(plot='\n'.join([html_table, compare_div])))
            else:
                print("did not make comparison plots for {}, nothing to compare".format(name))
            # dendro plot
            #fix_name_compare = fix_name + '_dendro'
            #make_dendogram_plot(name, set, loci_list, loci_dict, filename.format(fix_name_compare))

        # summerysing the results
        #r_div = make_readdepth_regression_plot(loci_dict)
        r_div2 = make_readdepth_regression_plot(loci_dict,False)
        filename = args.plots + 'sample_{}.html'
        with open(filename.format('summery'), 'w') as outfile:
            outfile.write(html_template.format(plot=r_div2))
    else:
        print("could not make plots, required files missing")

if __name__ == '__main__':
    args = argument_parse()
    main(args)