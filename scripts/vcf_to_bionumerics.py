#!/usr/local/bin/python3.6
"""
read a vcf file and accompanying get the allel calls for the samples
write the allel calls in a format that can be imported into biometrics.
"""

from openpyxl import load_workbook
from argparse import ArgumentParser, FileType
import sys
import vcf


def argument_parse():
    parser = ArgumentParser(description='Options for BarcodeSplit script')
    parser.add_argument(
        "-v", "--vcf",
        dest="vcf_file",
        required=True,
        help="the path to the vcf file containing varients of the GT-seq run."
    )
    parser.add_argument(
        '-i', '--info',
        required=True,
        help="the .csv or .xlsx file with information on the read groups"
    )
    parser.add_argument(
        '-o', '--outfile',
        nargs='?',
        type=FileType('w'),
        default = sys.stdout
    )
    return parser.parse_args()


def obtain_barcode_info_xlsx(info_file):
    info_lines = []
    wb = load_workbook(info_file, data_only=True)
    if 'Library' in wb.sheetnames:
        info_sheet = wb['Library']
    else:
        info_sheet = wb.active
    info_headers = []
    for column in info_sheet.iter_cols(max_row=1):
        info_headers.append(column[0].value)
    for row in info_sheet.iter_rows(min_row=2):
        info_dict = {info_headers[i]:item.value for i,item in enumerate(row)}
        # rows with empty cells are not allowed
        if all(x for x in info_dict.values()):
            info_lines.append(info_dict)
    return info_lines


def make_sample_list(rg_info_list, species=None):
    sample_list = {'keys':['Box-Positie','sample_name']}
    for sample in rg_info_list:
        if not species or sample['Species'] == species:
            sample_id = sample['Sample_name']
            sample_lable = sample['Plate_name']+'_'+sample['Position']
            sample_list[sample_id] = {'Box-Positie':sample_lable,'sample_name':sample['Sample_name']}
    return sample_list


def write_genotypes(vcf_reader,output, sample_list=None, gt_base_replace={None:'?'}):
    # because the output has the locus on the columns we would have to move trough the vcf for every line.
    # Better to obtain the required information in a dictionary in one go.
    loci_id_list = []
    genotype_dict = {sample: {} for sample in vcf_reader.samples}
    for loci in vcf_reader:
        # only add SNP that are not filtered.
        if loci.FILTER == []:
            # please do not change locus id, must
            site_id = loci.CHROM + '_{:07}'.format(loci.POS)
            loci_id_list.append(site_id)
            for sample in loci.samples:
                if sample.gt_bases in gt_base_replace:
                    genotype_dict[sample.sample][site_id] = gt_base_replace[sample.gt_bases]
                else:
                    genotype_dict[sample.sample][site_id] = sample.gt_bases
    #write the header
    if not sample_list:
        sample_list = {sample:{'sample':sample} for sample in vcf_reader.samples}
        output.write("sample;" + ';'.join(loci_id_list) + '\n')
        title = ['sample']
    else:
        title = sample_list.pop('keys')
        output.write(";".join(title) + ";" + ';'.join(loci_id_list) + '\n')
    #write the lines
    for sample,info in sample_list.items():
        output.write(';'.join([info[i] for i in title])+';'+';'.join([genotype_dict[sample][loci] for loci in loci_id_list])+'\n')


def main(args):
    # opening the vcf file
    vcf_reader = vcf.Reader(filename=args.vcf_file)
    rg_info = obtain_barcode_info_xlsx(args.info)
    bases = ['A','C','G','T']
    #gt_base_replace1 = {x+'/'+y:x+':'+y for x in bases for y in bases}
    #gt_base_replace1[None] = '?'
    gt_base_replace2 = {x+'/'+y:'5' for x in bases for y in bases if x!=y}
    gt_base_replace2[None] = '?'
    gt_base_replace2['A/A'] = '1'
    gt_base_replace2['C/C'] = '2'
    gt_base_replace2['G/G'] = '3'
    gt_base_replace2['T/T'] = '4'
    sample_list = make_sample_list(rg_info, species='Komkommer')
    write_genotypes(vcf_reader, args.outfile, gt_base_replace=gt_base_replace2, sample_list=sample_list)


if __name__ == '__main__':
    args = argument_parse()
    main(args)